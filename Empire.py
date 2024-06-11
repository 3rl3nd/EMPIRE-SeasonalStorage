from __future__ import division
from pyomo.environ import *
from pyomo.common.tempfiles import TempfileManager
import csv
import sys
import cloudpickle
import time
import os
from datetime import datetime
from openpyxl import load_workbook
import pandas as pd

# import cartopy
# import cartopy.crs as ccrs

# import matplotlib.pyplot as plt
# from matplotlib.lines import Line2D

def strfdelta(tdelta, fmt):
    d = {"days": tdelta.days}
    d["H"], rem = divmod(tdelta.seconds, 3600)
    d["H"] = str("{:02d}".format(d["H"]))
    d["M"], d["S"] = divmod(rem, 60)
    return fmt.format(**d)

# noinspection PyTypeChecker
def run_empire(name, tab_file_path, data_handler_path, result_file_path, branch_generation, branch_data_path,
               solver, temp_dir, FirstHoursOfRegSeason, lengthRegSeason,
               Period, NoOfPeriods, CurrentPeriods, last_run, updatePeriods, Operationalhour, Branch, ParentDictionary, ProbabilityDictionary,
               HoursAndSeasonOfBranch,HoursOfBranch, BranchesOfSeason, BranchPath, Season, HoursOfSeason, NoOfRegSeason,
               discountrate, WACC, LeapYearsInvestment, WRITE_LP, PICKLE_INSTANCE, EMISSION_CAP,include_results,
               USE_TEMP_DIR, MaxChargeAndDischargePercentage, SEASONAL_STORAGE, FLEX_HYDROGEN, HYDROGEN_CONSTANT_DEMAND, GREEN_HYDROGEN, REFORMER_HYDROGEN,
               RENEWABLE_GRID_RULE, start_year, H2LoadScale):

    if USE_TEMP_DIR:
        TempfileManager.tempdir = temp_dir

    if not os.path.exists(result_file_path):
        os.makedirs(result_file_path)

    GJperMWh = 3.6
    ng_MWhPerTon = 13.9
    hydrogen_MWhPerTon = 33
    coal_lhv_mj_per_kg = 29.0 # MJ/kg = GJ/ton

    co2_scale_factor = 1

    model = AbstractModel()

    ###########
    ##SOLVERS##
    ###########

    if solver == "CPLEX":
        print("Solver: CPLEX")
    elif solver == "Xpress":
        print("Solver: Xpress")
    elif solver == "Gurobi":
        print("Solver: Gurobi")
    else:
        sys.exit("ERROR! Invalid solver! Options: CPLEX, Xpress, Gurobi")

    ##########
    ##MODULE##
    ##########

    if WRITE_LP:
        print("Will write LP-file...")

    if PICKLE_INSTANCE:
        print("Will pickle instance...")

    if EMISSION_CAP:
        print("Absolute emission cap in each scenario...")
    else:
        print("No absolute emission cap...")

    print("Optimizing with hydrogen component")

    ########
    ##SETS##
    ########

    #Define the sets
    timeStart = datetime.now()
    print("Declaring sets...")

    #Supply technology sets
    model.Generator = Set(ordered=True) #g
    model.HydrogenGenerators = Set(ordered=True, within=model.Generator)
    model.Technology = Set(ordered=True) #t
    model.Storage = Set() #b

    #Temporal sets
    model.Period = Set(ordered=True, initialize=Period) #i
    model.CurrentPeriods = Set(ordered=True, initialize=CurrentPeriods) #i
    model.Operationalhour = Set(ordered=True, initialize=Operationalhour) #h
    model.Season = Set(ordered=True, initialize=Season) #s

    #Spatial sets
    model.Node = Set(ordered=True) #n
    # model.Node = Set(within=model.Node, ordered=True)
    model.DirectionalLink = Set(dimen=2, within=model.Node*model.Node, ordered=True) #a
    model.HydrogenBidirectionalLink = Set(dimen=2, within=model.Node*model.Node, ordered=True) #a
    model.TransmissionType = Set(ordered=True)
    
    #model.NaturalGasNode = Set(within=model.Node, ordered=True) #n  

    #Stochastic sets
    model.Branch = Set(ordered=True, initialize=Branch) #w
    model.HoursOfBranch = Set(dimen=2, ordered=True, initialize=HoursOfBranch)
    model.HoursAndSeasonOfBranch = Set(dimen=3, initialize=HoursAndSeasonOfBranch)
    model.BranchesOfSeason = Set(dimen=2, initialize=BranchesOfSeason)
    model.BranchPath = Set(dimen=12, initialize=BranchPath)
    model.ParentDictionary = ParentDictionary
    model.ProbabilityDictionary = ProbabilityDictionary

    #Subsets
    model.GeneratorsOfTechnology = Set(dimen=2) #(t,g) for all t in T, g in G_t
    model.GeneratorsOfNode = Set(dimen=2) #(n,g) for all n in N, g in G_n
    model.TransmissionTypeOfDirectionalLink = Set(dimen=3) #(n1,n2,t) for all (n1,n2) in L, t in T
    model.RampingGenerators = Set(within=model.Generator) #g_ramp
    model.RegHydroGenerator = Set(within=model.Generator) #g_reghyd
    model.HydroGenerator = Set(within=model.Generator) #g_hyd
    model.StoragesOfNode = Set(dimen=2) #(n,b) for all n in N, b in B_n
    model.DependentStorage = Set() #b_dagger
    model.HoursOfSeason = Set(dimen=2, ordered=True, initialize=HoursOfSeason) #(s,h) for all s in S, h in H_s
    model.FirstHoursOfRegSeason = Set(within=model.Operationalhour, ordered=True, initialize=FirstHoursOfRegSeason)

    print("Reading sets...")

    #Load the data

    data = DataPortal()
    data.load(filename=tab_file_path + "/" + 'Sets_Generator.tab',format="set", set=model.Generator)
    data.load(filename=tab_file_path + "/" + 'Sets_RampingGenerators.tab',format="set", set=model.RampingGenerators)
    data.load(filename=tab_file_path + "/" + 'Sets_HydroGenerator.tab',format="set", set=model.HydroGenerator)
    data.load(filename=tab_file_path + "/" + 'Sets_HydroGeneratorWithReservoir.tab',format="set", set=model.RegHydroGenerator)
    data.load(filename=tab_file_path + "/" + 'Sets_Storage.tab',format="set", set=model.Storage)
    data.load(filename=tab_file_path + "/" + 'Sets_DependentStorage.tab',format="set", set=model.DependentStorage)
    data.load(filename=tab_file_path + "/" + 'Sets_Technology.tab',format="set", set=model.Technology)
    data.load(filename=tab_file_path + "/" + 'Sets_Node.tab',format="set", set=model.Node)
    data.load(filename=tab_file_path + "/" + 'Sets_DirectionalLines.tab',format="set", set=model.DirectionalLink)
    data.load(filename=tab_file_path + "/" + 'Sets_LineType.tab',format="set", set=model.TransmissionType)
    data.load(filename=tab_file_path + "/" + 'Sets_LineTypeOfDirectionalLines.tab',format="set", set=model.TransmissionTypeOfDirectionalLink)
    data.load(filename=tab_file_path + "/" + 'Sets_GeneratorsOfTechnology.tab',format="set", set=model.GeneratorsOfTechnology)
    data.load(filename=tab_file_path + "/" + 'Sets_GeneratorsOfNode.tab',format="set", set=model.GeneratorsOfNode)
    data.load(filename=tab_file_path + "/" + 'Sets_StorageOfNodes.tab',format="set", set=model.StoragesOfNode)
    data.load(filename=tab_file_path + "/" + 'Sets_HydrogenLines.tab',format="set", set=model.HydrogenBidirectionalLink)


    print("Constructing sub sets...")

    #Build arc subsets

    def NodesLinked_init(model, node):
        retval = []
        for (i,j) in model.DirectionalLink:
            if j == node:
                retval.append(i)
        return retval
    model.NodesLinked = Set(model.Node, initialize=NodesLinked_init)

    def BidirectionalArc_init(model):
        retval = []
        for (i,j) in model.DirectionalLink:
            if (not (i == 'Ireland' and j == 'Spain')) or (not (j == 'Ireland' and i == 'Spain')):
                if i != j and (not (j,i) in retval):
                    retval.append((i,j))
        return retval
    model.BidirectionalArc = Set(dimen=2, initialize=BidirectionalArc_init, ordered=True) #l

    def prep_hydrogenGenerators_rule(model):
        for g in model.Generator:
                if "hydrogen" in g.lower():
                    model.HydrogenGenerators.add(g)
    model.build_hydrogenGenerators = BuildAction(rule=prep_hydrogenGenerators_rule)


    def NaturalGasGenerators_init(model):
        retval = []
        for gen in model.Generator:
            if 'gas' in gen.lower():
                retval.append(gen)
        return retval
    model.NaturalGasGenerators = Set(ordered=True, initialize=NaturalGasGenerators_init, within=model.Generator)

    ##############
    ##PARAMETERS##
    ##############

    #Define the parameters

    print("Declaring parameters...")

    #Scaling

    model.discountrate = Param(initialize=discountrate)
    model.WACC = Param(initialize=WACC)
    model.LeapYearsInvestment = Param(initialize=LeapYearsInvestment)
    model.operationalDiscountrate = Param(mutable=True)
    model.branchProbab = Param(model.Branch, mutable=True)
    model.seasScale = Param(model.Season, initialize=1.0, mutable=True)
    model.lengthRegSeason = Param(initialize=lengthRegSeason)

    #Cost

    model.genCapitalCost = Param(model.Generator, model.Period, default=0, mutable=True)
    model.transmissionTypeCapitalCost = Param(model.TransmissionType, model.Period, default=0, mutable=True)
    model.storPWCapitalCost = Param(model.Storage, model.Period, default=0, mutable=True)
    model.storENCapitalCost = Param(model.Storage, model.Period, default=0, mutable=True)
    model.genFixedOMCost = Param(model.Generator, model.Period, default=0, mutable=True)
    model.transmissionTypeFixedOMCost = Param(model.TransmissionType, model.Period, default=0, mutable=True)
    model.storPWFixedOMCost = Param(model.Storage, model.Period, default=0, mutable=True)
    model.storENFixedOMCost = Param(model.Storage, model.Period, default=0, mutable=True)
    model.genInvCost = Param(model.Generator, model.Period, default=9000000, mutable=True)
    model.transmissionInvCost = Param(model.BidirectionalArc, model.Period, default=3000000, mutable=True)
    model.storPWInvCost = Param(model.Storage, model.Period, default=1000000, mutable=True)
    model.storENInvCost = Param(model.Storage, model.Period, default=800000, mutable=True)
    model.transmissionLength = Param(model.BidirectionalArc, mutable=True)
    model.genVariableOMCost = Param(model.Generator, default=0.0, mutable=True)
    model.genFuelCost = Param(model.Generator, model.Period, mutable=True)
    model.genMargCost = Param(model.Generator, model.Period, default=600, mutable=True)
    model.genCO2TypeFactor = Param(model.Generator, default=0.0, mutable=True)
    model.genCO2Captured = Param(model.Generator, default=0.0, mutable=True)
    model.nodeLostLoadCost = Param(model.Node, model.Period, default=22000.0)
    model.CO2price = Param(model.Period, default=0.0, mutable=True)
    # model.CCSCostTSFix = Param(initialize=1149873.72) #NB! Hard-coded
    model.CCSCostTSVariable = Param(model.Period, default=0.0, mutable=True)
    # model.CCSRemFrac = Param(initialize=0.9)

    #Node dependent technology limitations

    model.genRefInitCap = Param(model.GeneratorsOfNode, default=0.0, mutable=True)
    model.genScaleInitCap = Param(model.Generator, model.Period, default=0.0, mutable=True)
    model.genInitCap = Param(model.GeneratorsOfNode, model.Period, default=0.0, mutable=True)
    model.transmissionInitCap = Param(model.BidirectionalArc, model.Period, default=0.0, mutable=True)
    model.storPWInitCap = Param(model.StoragesOfNode, model.Period, default=0.0, mutable=True)
    model.storENInitCap = Param(model.StoragesOfNode, model.Period, default=0.0, mutable=True)
    model.genMaxBuiltCap = Param(model.Node, model.Technology, model.Period, default=500000.0, mutable=True)
    model.transmissionMaxBuiltCap = Param(model.BidirectionalArc, model.Period, default=10000.0, mutable=True)
    model.storPWMaxBuiltCap = Param(model.StoragesOfNode, model.Period, default=500000.0, mutable=True)
    model.storENMaxBuiltCap = Param(model.StoragesOfNode, model.Period, default=500000.0, mutable=True)
    model.genMaxInstalledCapRaw = Param(model.Node, model.Technology, default=0.0, mutable=True)
    model.genMaxInstalledCap = Param(model.Node, model.Technology, model.Period, default=0.0, mutable=True)
    model.transmissionMaxInstalledCapRaw = Param(model.BidirectionalArc, model.Period, default=0.0)
    model.transmissionMaxInstalledCap = Param(model.BidirectionalArc, model.Period, default=0.0, mutable=True)
    model.storPWMaxInstalledCap = Param(model.StoragesOfNode, model.Period, default=0.0, mutable=True)
    model.storPWMaxInstalledCapRaw = Param(model.StoragesOfNode, default=0.0, mutable=True)
    model.storENMaxInstalledCap = Param(model.StoragesOfNode, model.Period, default=0.0, mutable=True)
    model.storENMaxInstalledCapRaw = Param(model.StoragesOfNode, default=0.0, mutable=True)

    model.genInitInv = Param(model.GeneratorsOfNode, model.Period, default=0.0, mutable=True)
    model.transInitInv = Param(model.BidirectionalArc, model.Period, default=0.0, mutable=True)
    model.storPWInitInv = Param(model.StoragesOfNode, model.Period, default=0.0, mutable=True)
    model.storENInitInv = Param(model.StoragesOfNode, model.Period, default=0.0, mutable=True)

    #Type dependent technology limitations

    model.genLifetime = Param(model.Generator, default=0.0, mutable=True)
    model.transmissionLifetime = Param(model.BidirectionalArc, default=40.0, mutable=True)
    model.storageLifetime = Param(model.Storage, default=0.0, mutable=True)
    model.genEfficiency = Param(model.Generator, model.Period, default=1.0, mutable=True)
    model.lineEfficiency = Param(model.DirectionalLink, default=0.97, mutable=True)
    model.storageChargeEff = Param(model.Storage, default=1.0, mutable=True)
    model.storageDischargeEff = Param(model.Storage, default=1.0, mutable=True)
    model.storageBleedEff = Param(model.Storage, default=1.0, mutable=True)
    model.genRampUpCap = Param(model.RampingGenerators, default=0.0, mutable=True)
    model.storageDiscToCharRatio = Param(model.Storage, default=1.0, mutable=True) #NB! Hard-coded
    model.storagePowToEnergy = Param(model.DependentStorage, default=1.0, mutable=True)

    #Stochastic input

    model.sloadRaw = Param(model.Node, model.Period, model.HoursOfBranch, default=0.0, mutable=True)
    model.sloadAnnualDemand = Param(model.Node, model.Period, default=0.0, mutable=True)
    model.sload = Param(model.Node, model.Period, model.HoursOfBranch, default=0.0, mutable=True)

    # h2load
    model.h2loadRaw = Param(model.Node, model.Period, model.HoursOfBranch, default=0.0, mutable=True)
    model.h2loadAnnualDemand = Param(model.Node, model.Period, default=0.0, mutable=True)
    model.h2load = Param(model.Node, model.Period, model.HoursOfBranch, default=0.0, mutable=True)
    
    model.genCapAvailTypeRaw = Param(model.Generator, default=1.0, mutable=True)
    model.genCapAvailStochRaw = Param(model.GeneratorsOfNode, model.HoursOfBranch, model.Period, default=0.0, mutable=True)
    model.genCapAvail = Param(model.GeneratorsOfNode, model.HoursOfBranch, model.Period, default=0.0, mutable=True)
    model.maxRegHydroGenRaw = Param(model.Node, model.Period, model.HoursAndSeasonOfBranch, default=1.0, mutable=True)
    model.maxRegHydroGen = Param(model.Node, model.Period, model.BranchesOfSeason, default=1.0, mutable=True)
    model.maxHydroNode = Param(model.Node, default=0.0, mutable=True)
    model.storOperationalInit = Param(model.Storage, default=0.0, mutable=True) #Percentage of installed energy capacity initially

    model.availableBioEnergy = Param(model.Period, default=0, mutable=True)

    if EMISSION_CAP:
        # co2_cap_exceeded_price = 10000
        # model.CO2CapExceeded = Var(model.Period, model.Branch, domain=NonNegativeReals)
        model.CO2cap = Param(model.Period, default=5000.0, mutable=True)

    #SÆVAREID: Coordinates for map visualization
    model.Latitude = Param(model.Node, default=0.0, mutable=True)
    model.Longitude = Param(model.Node, default=0.0, mutable=True)

    model.ng_cost = Param(model.Period, default=0, mutable=True)

    model.hydrogen_demand = Param(model.Node, model.Period)


    #Load the parameters

    print("Reading parameters...")

    data.load(filename=tab_file_path + "/" + 'Generator_CapitalCosts.tab', param=model.genCapitalCost, format="table")
    data.load(filename=tab_file_path + "/" + 'Generator_FixedOMCosts.tab', param=model.genFixedOMCost, format="table")
    data.load(filename=tab_file_path + "/" + 'Generator_VariableOMCosts.tab', param=model.genVariableOMCost, format="table")
    data.load(filename=tab_file_path + "/" + 'Generator_FuelCosts.tab', param=model.genFuelCost, format="table")
    data.load(filename=tab_file_path + "/" + 'Generator_CCSCostTSVariable.tab', param=model.CCSCostTSVariable, format="table")
    data.load(filename=tab_file_path + "/" + 'Generator_Efficiency.tab', param=model.genEfficiency, format="table")
    data.load(filename=tab_file_path + "/" + 'Generator_RefInitialCap.tab', param=model.genRefInitCap, format="table")
    data.load(filename=tab_file_path + "/" + 'Generator_ScaleFactorInitialCap.tab', param=model.genScaleInitCap, format="table")
    # data.load(filename=tab_file_path + "/" + 'Generator_InitialCapacity.tab', param=model.genInitCap, format="table") #node_generator_intial_capacity.xlsx
    data.load(filename=tab_file_path + "/" + 'Generator_MaxBuiltCapacity.tab', param=model.genMaxBuiltCap, format="table")#?
    data.load(filename=tab_file_path + "/" + 'Generator_MaxInstalledCapacity.tab', param=model.genMaxInstalledCapRaw, format="table")#maximum_capacity_constraint_040317_high
    data.load(filename=tab_file_path + "/" + 'Generator_CO2Content.tab', param=model.genCO2TypeFactor, format="table")
    data.load(filename=tab_file_path + "/" + 'Generator_CO2Captured.tab', param=model.genCO2Captured, format="table")
    data.load(filename=tab_file_path + "/" + 'Generator_RampRate.tab', param=model.genRampUpCap, format="table")
    data.load(filename=tab_file_path + "/" + 'Generator_GeneratorTypeAvailability.tab', param=model.genCapAvailTypeRaw, format="table")
    data.load(filename=tab_file_path + "/" + 'Generator_Lifetime.tab', param=model.genLifetime, format="table")
    data.load(filename=tab_file_path + "/" + 'Generator_GeneratorInv.tab', param=model.genInitInv, format="table")

    data.load(filename=tab_file_path + "/" + 'Transmission_InitialCapacity.tab', param=model.transmissionInitCap, format="table")
    data.load(filename=tab_file_path + "/" + 'Transmission_MaxBuiltCapacity.tab', param=model.transmissionMaxBuiltCap, format="table")
    data.load(filename=tab_file_path + "/" + 'Transmission_MaxInstallCapacityRaw.tab', param=model.transmissionMaxInstalledCapRaw, format="table")
    data.load(filename=tab_file_path + "/" + 'Transmission_Length.tab', param=model.transmissionLength, format="table")
    data.load(filename=tab_file_path + "/" + 'Transmission_TypeCapitalCost.tab', param=model.transmissionTypeCapitalCost, format="table")
    data.load(filename=tab_file_path + "/" + 'Transmission_TypeFixedOMCost.tab', param=model.transmissionTypeFixedOMCost, format="table")
    data.load(filename=tab_file_path + "/" + 'Transmission_lineEfficiency.tab', param=model.lineEfficiency, format="table")
    data.load(filename=tab_file_path + "/" + 'Transmission_Lifetime.tab', param=model.transmissionLifetime, format="table")
    data.load(filename=tab_file_path + "/" + 'Transmission_TransmissionInv.tab', param=model.transInitInv, format="table")

    data.load(filename=tab_file_path + "/" + 'Storage_StorageBleedEfficiency.tab', param=model.storageBleedEff, format="table")
    data.load(filename=tab_file_path + "/" + 'Storage_StorageChargeEff.tab', param=model.storageChargeEff, format="table")
    data.load(filename=tab_file_path + "/" + 'Storage_StorageDischargeEff.tab', param=model.storageDischargeEff, format="table")
    data.load(filename=tab_file_path + "/" + 'Storage_StoragePowToEnergy.tab', param=model.storagePowToEnergy, format="table")
    data.load(filename=tab_file_path + "/" + 'Storage_EnergyCapitalCost.tab', param=model.storENCapitalCost, format="table")
    data.load(filename=tab_file_path + "/" + 'Storage_EnergyFixedOMCost.tab', param=model.storENFixedOMCost, format="table")
    data.load(filename=tab_file_path + "/" + 'Storage_EnergyInitialCapacity.tab', param=model.storENInitCap, format="table")
    data.load(filename=tab_file_path + "/" + 'Storage_EnergyMaxBuiltCapacity.tab', param=model.storENMaxBuiltCap, format="table")
    data.load(filename=tab_file_path + "/" + 'Storage_EnergyMaxInstalledCapacity.tab', param=model.storENMaxInstalledCapRaw, format="table")
    data.load(filename=tab_file_path + "/" + 'Storage_StorageInitialEnergyLevel.tab', param=model.storOperationalInit, format="table")
    data.load(filename=tab_file_path + "/" + 'Storage_PowerCapitalCost.tab', param=model.storPWCapitalCost, format="table")
    data.load(filename=tab_file_path + "/" + 'Storage_PowerFixedOMCost.tab', param=model.storPWFixedOMCost, format="table")
    data.load(filename=tab_file_path + "/" + 'Storage_InitialPowerCapacity.tab', param=model.storPWInitCap, format="table")
    data.load(filename=tab_file_path + "/" + 'Storage_PowerMaxBuiltCapacity.tab', param=model.storPWMaxBuiltCap, format="table")
    data.load(filename=tab_file_path + "/" + 'Storage_PowerMaxInstalledCapacity.tab', param=model.storPWMaxInstalledCapRaw, format="table")
    data.load(filename=tab_file_path + "/" + 'Storage_Lifetime.tab', param=model.storageLifetime, format="table")
    data.load(filename=tab_file_path + "/" + 'Storage_StoragePWInv.tab', param=model.storPWInitInv, format="table")
    data.load(filename=tab_file_path + "/" + 'Storage_StorageENInv.tab', param=model.storENInitInv, format="table")

    data.load(filename=tab_file_path + "/" + 'Node_NodeLostLoadCost.tab', param=model.nodeLostLoadCost, format="table")
    data.load(filename=tab_file_path + "/" + 'Node_ElectricAnnualDemand.tab', param=model.sloadAnnualDemand, format="table")
    # h2load
    data.load(filename=tab_file_path + "/" + 'Node_HydrogenAnnualDemand.tab', param=model.h2loadAnnualDemand, format="table")
    data.load(filename=tab_file_path + "/" + 'Node_HydroGenMaxAnnualProduction.tab', param=model.maxHydroNode, format="table")

    #SÆVAREID: Coordinates
    data.load(filename=tab_file_path + "/" + 'Node_Latitude.tab', param=model.Latitude, format="table")
    data.load(filename=tab_file_path + "/" + 'Node_Longitude.tab', param=model.Longitude, format="table")

    data.load(filename=tab_file_path + '/' + 'NaturalGas_PeriodCost.tab',param=model.ng_cost, format='table') # We need the cost per period 


    if branch_generation:
        branch_file_path = tab_file_path
    else:
        branch_file_path = branch_data_path

    data.load(filename=branch_file_path + "/" + 'Stochastic_HydroGenMaxSeasonalProduction.tab', param=model.maxRegHydroGenRaw, format="table")
    data.load(filename=branch_file_path + "/" + 'Stochastic_StochasticAvailability.tab', param=model.genCapAvailStochRaw, format="table")
    data.load(filename=branch_file_path + "/" + 'Stochastic_ElectricLoadRaw.tab', param=model.sloadRaw, format="table")
    # h2load
    data.load(filename=branch_file_path + "/" + 'Stochastic_HydrogenLoadRaw.tab', param=model.h2loadRaw, format="table")

    # data.load(filename=tab_file_path + "/" + 'General_seasonScale.tab', param=model.seasScale, format="table")
    data.load(filename=tab_file_path + "/" + 'General_AvailableBioEnergy.tab', param=model.availableBioEnergy, format="table")

    data.load(filename=tab_file_path + "/" + 'General_CO2Cap.tab', param=model.CO2cap, format="table")
    data.load(filename=tab_file_path + "/" + 'General_CO2Price.tab', param=model.CO2price, format="table")

    print("Constructing parameter values...")

    def prepbranchProbab_rule(model):
        #Build an equiprobable probability distribution for scenarios

        for w in model.Branch:
            model.branchProbab[w] = value(model.ProbabilityDictionary[w])

    model.build_branchProbab = BuildAction(rule=prepbranchProbab_rule)

    def prepSeasScale(model):
        for s in model.Season:
            model.seasScale[s] = 8760/(lengthRegSeason*NoOfRegSeason)  
    model.build_seasScale = BuildAction(rule=prepSeasScale)

    def prepInvCost_rule(model):
        #Build investment cost for generators, storages and transmission. Annual cost is calculated for the lifetime of the generator and discounted for a year.
        #Then cost is discounted for the investment period (or the remaining lifetime). CCS generators has additional fixed costs depending on emissions.
        #Generator
        for g in model.Generator:
            for i in model.CurrentPeriods:
                costperyear=(model.WACC / (1 - ((1+model.WACC) ** (1-model.genLifetime[g])))) * model.genCapitalCost[g,i] + model.genFixedOMCost[g,i]
                costperperiod = costperyear * 1000 * (1 - (1+model.discountrate) **-(min((NoOfPeriods-i+1)*value(model.LeapYearsInvestment), value(model.genLifetime[g]))))/ (1 - (1 / (1 + model.discountrate)))
                # if ('CCS',g) in model.GeneratorsOfTechnology:
                    # 	costperperiod+=model.CCSCostTSFix*model.CCSRemFrac*model.genCO2TypeFactor[g]*(GJperMWh/model.genEfficiency[g,i])
                model.genInvCost[g,i]=costperperiod

        #Storage
        for b in model.Storage:
            for i in model.CurrentPeriods:
                costperyearPW=(model.WACC/(1-((1+model.WACC)**(1-model.storageLifetime[b]))))*model.storPWCapitalCost[b,i]+model.storPWFixedOMCost[b,i]
                costperperiodPW=costperyearPW*1000*(1-(1+model.discountrate)**-(min((NoOfPeriods-i+1)*value(model.LeapYearsInvestment), value(model.storageLifetime[b]))))/(1-(1/(1+model.discountrate)))
                model.storPWInvCost[b,i]=costperperiodPW
                costperyearEN=(model.WACC/(1-((1+model.WACC)**(1-model.storageLifetime[b]))))*model.storENCapitalCost[b,i]+model.storENFixedOMCost[b,i]
                costperperiodEN=costperyearEN*1000*(1-(1+model.discountrate)**-(min((NoOfPeriods-i+1)*value(model.LeapYearsInvestment), value(model.storageLifetime[b]))))/(1-(1/(1+model.discountrate)))
                model.storENInvCost[b,i]=costperperiodEN

        #Transmission
        for (n1,n2) in model.BidirectionalArc:
            for i in model.CurrentPeriods:
                for t in model.TransmissionType:
                    if (n1,n2,t) in model.TransmissionTypeOfDirectionalLink:
                        costperyear=(model.WACC/(1-((1+model.WACC)**(1-model.transmissionLifetime[n1,n2]))))*model.transmissionLength[n1,n2]*model.transmissionTypeCapitalCost[t,i] + model.transmissionLength[n1,n2]* model.transmissionTypeFixedOMCost[t,i]
                        costperperiod=costperyear*(1-(1+model.discountrate)**-(min((NoOfPeriods-i+1)*value(model.LeapYearsInvestment), value(model.transmissionLifetime[n1,n2]))))/(1-(1/(1+model.discountrate)))
                        model.transmissionInvCost[n1,n2,i]=costperperiod

    model.build_InvCost = BuildAction(rule=prepInvCost_rule)
    

    def prepOperationalCostGen_rule(model):
        for g in model.Generator:
            for i in model.CurrentPeriods:
                if not EMISSION_CAP:
                    costperenergyunit=(GJperMWh/model.genEfficiency[g,i])*(model.genCO2TypeFactor[g]*model.CO2price[i])+ \
                                    model.genVariableOMCost[g]
                    if g not in model.NaturalGasGenerators and g not in model.HydrogenGenerators:
                        costperenergyunit += (GJperMWh/model.genEfficiency[g,i])*(model.genFuelCost[g,i])
                    if g in model.NaturalGasGenerators:
                        costperenergyunit += (model.ng_cost[i])/(ng_MWhPerTon*model.genEfficiency[g,i])
                else:
                    costperenergyunit = model.genVariableOMCost[g]
                    if g not in model.NaturalGasGenerators and g not in model.HydrogenGenerators:
                        costperenergyunit += (GJperMWh/model.genEfficiency[g,i])*(model.genFuelCost[g,i])
                    if g in model.NaturalGasGenerators:
                        costperenergyunit += (model.ng_cost[i])/(ng_MWhPerTon*model.genEfficiency[g,i]) # NG cost given per ton as input. To get cost per power output, divide this by 
                model.genMargCost[g,i] = costperenergyunit

        # OLD:
        # for g in model.Generator:
        #     for i in model.CurrentPeriods:
        #         if not EMISSION_CAP:
        #             if ('CCS',g) in model.GeneratorsOfTechnology:
        #                 costperenergyunit=(GJperMWh/model.genEfficiency[g,i])*(model.genCO2TypeFactor[g]*model.CO2price[i])+ \
        #                                   model.genVariableOMCost[g]
        #                 if g not in model.NaturalGasGenerators and g not in model.HydrogenGenerators:
        #                     costperenergyunit += (GJperMWh/model.genEfficiency[g,i])*(model.genFuelCost[g,i])
        #             else:
        #                 costperenergyunit=(GJperMWh/model.genEfficiency[g,i])*(model.genCO2TypeFactor[g]*model.CO2price[i])+ \
        #                                   model.genVariableOMCost[g]
        #                 if g not in model.NaturalGasGenerators and g not in model.HydrogenGenerators:
        #                     costperenergyunit += (GJperMWh/model.genEfficiency[g,i])*(model.genFuelCost[g,i])
        #         else:
        #             if ('CCS',g) in model.GeneratorsOfTechnology:
        #                 costperenergyunit = model.genVariableOMCost[g]
        #                 if g not in model.NaturalGasGenerators and g not in model.HydrogenGenerators:
        #                     costperenergyunit += (GJperMWh/model.genEfficiency[g,i])*(model.genFuelCost[g,i])
        #             else:
        #                 costperenergyunit= model.genVariableOMCost[g]
        #                 if g not in model.NaturalGasGenerators and g not in model.HydrogenGenerators:
        #                     costperenergyunit += (GJperMWh/model.genEfficiency[g,i])*(model.genFuelCost[g,i])
        #         model.genMargCost[g,i]=costperenergyunit

    model.build_OperationalCostGen = BuildAction(rule=prepOperationalCostGen_rule)

    def prepInitialCapacityNodeGen_rule(model):
        #Build initial capacity for generator type in node

        for (n,g) in model.GeneratorsOfNode:
            for i in model.CurrentPeriods:
                # if value(model.genInitCap[n,g,i]) == 0:
                model.genInitCap[n,g,i] = model.genRefInitCap[n,g]*(1-model.genScaleInitCap[g,i])

    model.build_InitialCapacityNodeGen = BuildAction(rule=prepInitialCapacityNodeGen_rule)

    def prepInitialCapacityTransmission_rule(model):
        #Build initial capacity for transmission lines to ensure initial capacity is the upper installation bound if infeasible

        for (n1,n2) in model.BidirectionalArc:
            for i in model.CurrentPeriods:
                if value(model.transmissionMaxInstalledCapRaw[n1,n2,i]) <= value(model.transmissionInitCap[n1,n2,i]):
                    model.transmissionMaxInstalledCap[n1,n2,i] = model.transmissionInitCap[n1,n2,i]
                else:
                    model.transmissionMaxInstalledCap[n1,n2,i] = model.transmissionMaxInstalledCapRaw[n1,n2,i]
    model.build_InitialCapacityTransmission = BuildAction(rule=prepInitialCapacityTransmission_rule)

    def prepOperationalDiscountrate_rule(model):
        #Build operational discount rate

        model.operationalDiscountrate = sum((1+model.discountrate)**(-j) for j in list(range(0,value(model.LeapYearsInvestment))))
    model.build_operationalDiscountrate = BuildAction(rule=prepOperationalDiscountrate_rule)

    def prepGenMaxInstalledCap_rule(model):
        #Build resource limit (installed limit) for all periods. Avoid infeasibility if installed limit lower than initially installed cap.

        for t in model.Technology:
            for n in model.Node:
                for i in model.CurrentPeriods:
                    if value(model.genMaxInstalledCapRaw[n,t] <= sum(model.genInitCap[n,g,i] for g in model.Generator if (n,g) in model.GeneratorsOfNode and (t,g) in model.GeneratorsOfTechnology)):
                        model.genMaxInstalledCap[n,t,i]=sum(model.genInitCap[n,g,i] for g in model.Generator if (n,g) in model.GeneratorsOfNode and (t,g) in model.GeneratorsOfTechnology)
                    else:
                        model.genMaxInstalledCap[n,t,i]=model.genMaxInstalledCapRaw[n,t]
    model.build_genMaxInstalledCap = BuildAction(rule=prepGenMaxInstalledCap_rule)

    def storENMaxInstalledCap_rule(model):
        #Build installed limit (resource limit) for storEN

        #Why is this here? Why not just use storENMaxInstalledCapRaw in the constraints?

        for (n,b) in model.StoragesOfNode:
            for i in model.CurrentPeriods:
                model.storENMaxInstalledCap[n,b,i]=model.storENMaxInstalledCapRaw[n,b]

    model.build_storENMaxInstalledCap = BuildAction(rule=storENMaxInstalledCap_rule)

    def storPWMaxInstalledCap_rule(model):
        #Build installed limit (resource limit) for storPW

        #Why is this here? Why not just use storPWMaxInstalledCapRaw in the constraints?

        for (n,b) in model.StoragesOfNode:
            for i in model.CurrentPeriods:
                model.storPWMaxInstalledCap[n,b,i]=model.storPWMaxInstalledCapRaw[n,b]

    model.build_storPWMaxInstalledCap = BuildAction(rule=storPWMaxInstalledCap_rule)

    def prepRegHydro_rule(model):
        #Build hydrolimits for all periods
        for n in model.Node:
            for (s,w) in model.BranchesOfSeason:
                for i in model.CurrentPeriods:
                    model.maxRegHydroGen[n,i,s,w]=sum(model.maxRegHydroGenRaw[n,i,w,s,h] for h in model.Operationalhour if (s,h) in model.HoursOfSeason)

    model.build_maxRegHydroGen = BuildAction(rule=prepRegHydro_rule)

    def prepGenCapAvail_rule(model):
        #Build generator availability for all periods

        for (n,g) in model.GeneratorsOfNode:
            for (w,h) in model.HoursOfBranch:
                    for i in model.CurrentPeriods:
                        if value(model.genCapAvailTypeRaw[g]) == 0:
                            if value(model.genCapAvailStochRaw[n,g,w,h,i]) >= 0.001:
                                model.genCapAvail[n,g,w,h,i] = model.genCapAvailStochRaw[n,g,w,h,i]
                            else:
                                model.genCapAvail[n,g,w,h,i] = 0
                        else:
                            model.genCapAvail[n,g,w,h,i]=model.genCapAvailTypeRaw[g]

    model.build_genCapAvail = BuildAction(rule=prepGenCapAvail_rule)

    def prepSload_rule(model):
        #Build load profiles for all periods

        counter = 0
        f = open(result_file_path + '/AdjustedNegativeLoad_' + name + '.txt', 'w')
        for n in model.Node:
            for i in model.CurrentPeriods:
                noderawdemand = value(sum(model.branchProbab[w]*model.seasScale[s]*model.sloadRaw[n,i,w,h] for (w,s,h) in model.HoursAndSeasonOfBranch))
                # nodeaverageload = nodeaverageload / value(
                # 	(model.FirstHoursOfRegSeason[-1] + model.lengthRegSeason - 1) * len(model.Branch))
                if noderawdemand > 0:
                    hourlyscale = model.sloadAnnualDemand[n,i].value / noderawdemand
                else:
                    hourlyscale = 0
                for (w,h) in model.HoursOfBranch:
                        model.sload[n, i, w, h] = model.sloadRaw[n,i,w,h]*hourlyscale
                        # if value(model.sloadRaw[n, h, w, i].value + hourlyadjustment) > 0:
                        # 	model.sload[n, h, i, w] = model.sloadRaw[n, h, w, i].value + hourlyadjustment
                        if value(model.sload[n,i,w,h]) < 0:
                            f.write('Adjusted electricity load: ' + str(value(model.sload[n,i,w,h])) + ', 10 MW for hour ' + str(h) + ' in period ' + str(i) + ' and branch ' + str(w) + ' in ' + str(n) + "\n")
                            model.sload[n,i,w,h] = 10
                            counter += 1
                        # else:
                        # 	f.write('Adjusted electricity load: ' + str(value(model.sloadRaw[n,h,sce,i].value + hourlyadjustment)) + ', 0 MW for hour ' + str(h) + ' and scenario ' + str(sce) + ' in ' + str(n) + "\n")
                        # 	model.sload[n,h,i,sce] = 0
                        # 	counter += 1
        f.write('Hours with too small raw electricity load: ' + str(counter))
        f.close()

    model.build_sload = BuildAction(rule=prepSload_rule)

    def prepH2load_rule(model):
        #Build load profiles for all periods

        counter = 0
        for n in model.Node:
            for i in model.CurrentPeriods:
                if HYDROGEN_CONSTANT_DEMAND:
                    for (w,h) in model.HoursOfBranch:
                        model.h2load[n, i, w, h] = value(model.h2loadAnnualDemand[n,i]*H2LoadScale) / 8760
                else:
                    noderawdemand = value(sum(model.branchProbab[w]*model.seasScale[s]*model.h2loadRaw[n,i,w,h] for (w,s,h) in model.HoursAndSeasonOfBranch))
                    if noderawdemand > 0:
                        hourlyscale = value(model.h2loadAnnualDemand[n,i]*H2LoadScale) / noderawdemand
                    else:
                        hourlyscale = 0
                    for (w,h) in model.HoursOfBranch:
                        model.h2load[n, i, w, h] = model.h2loadRaw[n,i,w,h]*hourlyscale

    model.build_H2load = BuildAction(rule=prepH2load_rule)

    stopReading = startConstraints = datetime.now()
    print("Sets and parameters declared and read...")

    #############
    ##VARIABLES##
    #############

    print("Declaring variables...")

    model.genInvCap = Var(model.GeneratorsOfNode, model.CurrentPeriods, domain=NonNegativeReals)
    model.transmissionInvCap = Var(model.BidirectionalArc, model.CurrentPeriods, domain=NonNegativeReals)
    model.storPWInvCap = Var(model.StoragesOfNode, model.CurrentPeriods, domain=NonNegativeReals)
    model.storENInvCap = Var(model.StoragesOfNode, model.CurrentPeriods, domain=NonNegativeReals)
    model.genOperational = Var(model.GeneratorsOfNode, model.CurrentPeriods, model.HoursOfBranch, domain=NonNegativeReals)
    model.storOperational = Var(model.StoragesOfNode, model.CurrentPeriods, model.HoursOfBranch, domain=NonNegativeReals)
    model.transmissionOperational = Var(model.DirectionalLink, model.CurrentPeriods, model.HoursOfBranch, domain=NonNegativeReals) #flow
    model.storCharge = Var(model.StoragesOfNode, model.CurrentPeriods, model.HoursOfBranch, domain=NonNegativeReals)
    model.storDischarge = Var(model.StoragesOfNode, model.CurrentPeriods, model.HoursOfBranch, domain=NonNegativeReals)
    model.loadShed = Var(model.Node, model.CurrentPeriods, model.HoursOfBranch, domain=NonNegativeReals)
    model.genInstalledCap = Var(model.GeneratorsOfNode, model.CurrentPeriods, domain=NonNegativeReals)
    model.transmissionInstalledCap = Var(model.BidirectionalArc, model.CurrentPeriods, domain=NonNegativeReals)
    model.storPWInstalledCap = Var(model.StoragesOfNode, model.CurrentPeriods, domain=NonNegativeReals)
    model.storENInstalledCap = Var(model.StoragesOfNode, model.CurrentPeriods, domain=NonNegativeReals)

    model.hydrogenDemandMet = Var(model.Node, model.CurrentPeriods, model.HoursOfBranch, within=NonNegativeReals)
    model.hydrogenDemandShed = Var(model.Node, model.CurrentPeriods, model.HoursOfBranch, within=NonNegativeReals)

    if REFORMER_HYDROGEN:
        model.ReformerPlants = Set(ordered=True)

    # if GREEN_HYDROGEN:
    model.RESGenerators = Set(within=model.Generator) # Subset for all renweable electricity generators
    if RENEWABLE_GRID_RULE:
        model.NonRenewableGrid = Set(dimen=2, within=model.Node*model.CurrentPeriods, ordered=True)
    
    if REFORMER_HYDROGEN:
        data.load(filename=tab_file_path + '/' + 'Hydrogen_ReformerPlants.tab', format="set", set=model.ReformerPlants)
    
    # if GREEN_HYDROGEN:
    data.load(filename=tab_file_path + '/' + 'Sets_RESGenerators.tab',format="set", set=model.RESGenerators)
    if RENEWABLE_GRID_RULE:
        data.load(filename=tab_file_path + '/' + 'Sets_NonRenewableGrid.tab',format="set", set=model.NonRenewableGrid)
    # data.load(filename=tab_file_path + '/' + 'Hydrogen_Generators.tab', format="set", set=model.HydrogenGenerators)

    def HydrogenLinks_init(model):
        retval= []
        for (n1,n2) in model.DirectionalLink:
            if (n1,n2) in model.HydrogenBidirectionalLink or (n2,n1) in model.HydrogenBidirectionalLink:
                retval.append((n1,n2))
        return retval
    model.AllowedHydrogenLinks = Set(dimen=2, initialize=HydrogenLinks_init, ordered=True)
    # model.AllowedHydrogenLinks = Set(dimen=2, within=model.Node * model.Node, ordered=True) # Depcreated; The links are now instead defined by the transmission links, but only between the production nodes
    # data.load(filename=tab_file_path + '/' + 'Hydrogen_Links.tab', format="set", set=model.AllowedHydrogenLinks) # Deprecated; The links are now instead defined by the transmission links, but only between the production nodes

    def HydrogenBidirectionPipelines_init(model):
        retval = []
        for (n1,n2) in model.BidirectionalArc:
            if (n1,n2) in model.HydrogenBidirectionalLink:
                retval.append((n1,n2))
        return retval
    model.HydrogenBidirectionPipelines = Set(dimen=2, initialize=HydrogenBidirectionPipelines_init, ordered=True)

    def HydrogenLinks_init(model, node):
        retval = []
        for (i,j) in model.AllowedHydrogenLinks:
            if j == node:
                retval.append(i)
        return retval
    model.HydrogenLinks = Set(model.Node, initialize=HydrogenLinks_init)

    model.elyzerPlantCapitalCost = Param(model.Period, default=99999, mutable=True)
    model.elyzerStackCapitalCost = Param(model.Period, default=99999, mutable=True)
    model.elyzerFixedOMCost = Param(model.Period, default=99999, mutable=True)
    model.elyzerPowerConsumptionPerTon = Param(model.Period, default=99999, mutable=True)
    model.elyzerLifetime = Param(default=20, mutable=True)
    model.elyzerInvCost = Param(model.Period, default=99999, mutable=True)

    if REFORMER_HYDROGEN:
        model.ReformerPlantsCapitalCost = Param(model.ReformerPlants, model.Period, default=99999, mutable=True)
        model.ReformerPlantFixedOMCost = Param(model.ReformerPlants, model.Period, default=99999, mutable=True)
        model.ReformerPlantVarOMCost = Param(model.ReformerPlants, model.Period, default=99999, mutable=True)
        model.ReformerPlantInvCost = Param(model.ReformerPlants, model.Period, default=99999, mutable=True)
        model.ReformerPlantEfficiency = Param(model.ReformerPlants, model.Period, default=0, mutable=True)
        model.ReformerFuelCost = Param(model.ReformerPlants, model.Period, default=0, mutable=True)
        model.ReformerPlantElectricityUse = Param(model.ReformerPlants, model.Period, default=99999, mutable=True)
        model.ReformerPlantLifetime = Param(model.ReformerPlants, default=25, mutable=True)
        model.ReformerEmissionFactor = Param(model.ReformerPlants, model.Period, default=99999, mutable=True)
        model.ReformerCO2CaptureFactor = Param(model.ReformerPlants, model.Period, default=99999, mutable=True)
        model.ReformerMargCost = Param(model.ReformerPlants, model.Period, default=99999, mutable=True)
        model.reformerInitInv = Param(model.Node ,model.ReformerPlants, model.Period, default=0, mutable=True)

    model.hydrogenPipelineLifetime = Param(default=40)
    model.hydrogenPipelineCapCost = Param(model.Period, default=99999, mutable=True)
    model.hydrogenPipelineOMCost = Param(model.Period, default=99999, mutable=True)
    model.hydrogenPipelineInvCost = Param(model.HydrogenBidirectionPipelines, model.Period, default=999999, mutable=True)
    model.PipelineLength = Param(model.HydrogenBidirectionPipelines, mutable=True, default=9999)
    model.hydrogenPipelineCompressorElectricityUsage = Param(default=99999, mutable=True)
    model.hydrogenPipelinePowerDemandPerTon = Param(model.HydrogenBidirectionPipelines, default=99999, mutable=True)

    # model.hydrogenPriceOtherMarkets = Param(default = h2priceOtherMarkets, mutable=True) #Price of 1 kg of H2 in other markets. This price is set by doing a sensitivity analysis and chosing a reasonable number

    # if h2storage is False:
    # 	#Cost of storing the produced hydrogen intraseasonally. Have to have this because we have implicit free storage without.
    # 	model.averageHydrogenSeasonalStorageCost = Param(default=0.35, mutable=True) #Source: Levelized cost of storage from Table 5 in Picturing the value of underground gas storage to the European hydrogen system by Gas Infrastructure Europe (GIE)
    model.hydrogenMaxStorageCapacity = Param(model.Node, default=0, mutable=True)
    model.hydrogenStorageCapitalCost = Param(model.Period, default=99999, mutable=True)
    model.hydrogenStorageFixedOMCost = Param(model.Period, default=99999, mutable=True)
    model.hydrogenStorageInvCost = Param(model.Period, default=99999, mutable=True)
    model.hydrogenStorageInitOperational = Param(default=0.5)
    model.hydrogenStorageLifetime = Param(default=30)

    model.hydrogenLHV_ton = Param(default=33.3, mutable=False) #LHV of hydrogen is 33.3 kWh/kg = 0.0333 MWh / kg = 33.3 MWh/ton
        
    model.elyzerInitInv = Param(model.Node, model.Period, default=0.0, mutable=True)
    model.hydrogenPipeInitInv = Param(model.BidirectionalArc, model.Period, default=0.0, mutable=True)
    model.hydrogenStorInitInv = Param(model.Node, model.Period, default=0.0, mutable=True)
    
    # if GREEN_HYDROGEN:
    model.RESGenInitInv = Param(model.Node, model.RESGenerators, model.Period, default=0.0, mutable=True)
    data.load(filename=tab_file_path + '/' + 'Generator_RESGeneratorInv.tab', format="table", param=model.RESGenInitInv)
    #Hydrogen variables
    #Operational
    model.hydrogenProducedElectro_ton = Var(model.Node, model.CurrentPeriods, model.HoursOfBranch, domain=NonNegativeReals)

    if REFORMER_HYDROGEN:
        model.hydrogenProducedReformer_ton = Var(model.Node, model.ReformerPlants, model.CurrentPeriods, model.HoursOfBranch, domain=NonNegativeReals)
        model.hydrogenProducedReformer_MWh = Var(model.Node, model.ReformerPlants, model.CurrentPeriods, model.HoursOfBranch, domain=NonNegativeReals)
        model.ng_forHydrogen = Var(model.Node, model.ReformerPlants, model.CurrentPeriods, model.HoursOfBranch, domain=NonNegativeReals)
        model.ReformerCapBuilt = Var(model.Node, model.ReformerPlants, model.CurrentPeriods, domain=NonNegativeReals) #Capacity  of MW H2 production built in period i
        model.ReformerTotalCap = Var(model.Node, model.ReformerPlants, model.CurrentPeriods, domain=NonNegativeReals) #Total capacity of MW H2 production


    model.hydrogenSentPipeline = Var(model.AllowedHydrogenLinks, model.CurrentPeriods, model.HoursOfBranch, domain=NonNegativeReals)
    model.powerForHydrogen = Var(model.Node, model.CurrentPeriods, model.HoursOfBranch, domain=NonNegativeReals) #Two period indexes because one describes the year it was bought (the first index), the other describes when it is used (second index)

    model.hydrogenStorageOperational = Var(model.Node, model.CurrentPeriods, model.HoursOfBranch, domain=NonNegativeReals)
    model.hydrogenChargeStorage = Var(model.Node, model.CurrentPeriods, model.HoursOfBranch, domain=NonNegativeReals, initialize=0)
    model.hydrogenDischargeStorage = Var(model.Node, model.CurrentPeriods, model.HoursOfBranch, domain=NonNegativeReals, initialize=0)

    model.hydrogenForPower = Var(model.HydrogenGenerators, model.Node, model.CurrentPeriods, model.HoursOfBranch, domain=NonNegativeReals,initialize=0.0)

    #Strategic
    model.elyzerCapBuilt = Var(model.Node, model.CurrentPeriods, domain=NonNegativeReals)
    model.elyzerTotalCap = Var(model.Node, model.CurrentPeriods, domain=NonNegativeReals)
    model.hydrogenPipelineBuilt = Var(model.HydrogenBidirectionPipelines, model.CurrentPeriods, domain=NonNegativeReals)
    model.totalHydrogenPipelineCapacity = Var(model.HydrogenBidirectionPipelines, model.CurrentPeriods, domain=NonNegativeReals)
    model.hydrogenStorageBuilt = Var(model.Node, model.CurrentPeriods, domain=NonNegativeReals)
    model.hydrogenTotalStorage = Var(model.Node, model.CurrentPeriods, domain=NonNegativeReals)

    # if GREEN_HYDROGEN:
    model.RESgenInstalledCap = Var(model.GeneratorsOfNode, model.CurrentPeriods, domain=NonNegativeReals)

    #Reading parameters
    data.load(filename=tab_file_path + '/' + 'Hydrogen_ElectrolyzerPlantCapitalCost.tab', format="table", param=model.elyzerPlantCapitalCost)
    data.load(filename=tab_file_path + '/' + 'Hydrogen_ElectrolyzerStackCapitalCost.tab', format="table", param=model.elyzerStackCapitalCost)
    data.load(filename=tab_file_path + '/' + 'Hydrogen_ElectrolyzerFixedOMCost.tab', format="table", param=model.elyzerFixedOMCost)
    data.load(filename=tab_file_path + '/' + 'Hydrogen_ElectrolyzerPowerUse.tab', format="table", param=model.elyzerPowerConsumptionPerTon)
    data.load(filename=tab_file_path + '/' + 'Hydrogen_ElectrolyzerLifetime.tab', format="table", param=model.elyzerLifetime)

    if REFORMER_HYDROGEN:
        data.load(filename=tab_file_path + '/' + 'Hydrogen_ReformerCapitalCost.tab', format='table', param=model.ReformerPlantsCapitalCost)
        data.load(filename=tab_file_path + '/' + 'Hydrogen_ReformerFixedOMCost.tab', format='table', param=model.ReformerPlantFixedOMCost)
        data.load(filename=tab_file_path + '/' + 'Hydrogen_ReformerVariableOMCost.tab', format='table', param=model.ReformerPlantVarOMCost)
        data.load(filename=tab_file_path + '/' + 'Hydrogen_ReformerEfficiency.tab', format='table', param=model.ReformerPlantEfficiency)
        data.load(filename=tab_file_path + '/' + 'Hydrogen_ReformerFuelCost.tab', format='table', param=model.ReformerFuelCost)
        data.load(filename=tab_file_path + '/' + 'Hydrogen_ReformerElectricityUse.tab', format='table', param=model.ReformerPlantElectricityUse)
        data.load(filename=tab_file_path + '/' + 'Hydrogen_ReformerLifetime.tab', format='table', param=model.ReformerPlantLifetime)
        data.load(filename=tab_file_path + '/' + 'Hydrogen_ReformerEmissionFactor.tab', format='table', param=model.ReformerEmissionFactor)
        data.load(filename=tab_file_path + '/' + 'Hydrogen_ReformerCO2CaptureFactor.tab', format='table', param=model.ReformerCO2CaptureFactor)
        data.load(filename=tab_file_path + '/' + 'Hydrogen_ReformerInv.tab', format="table", param=model.reformerInitInv)

    data.load(filename=tab_file_path + '/' + 'Hydrogen_PipelineCapitalCost.tab', format="table", param=model.hydrogenPipelineCapCost)
    data.load(filename=tab_file_path + '/' + 'Hydrogen_PipelineOMCostPerKM.tab', format="table", param=model.hydrogenPipelineOMCost)
    data.load(filename=tab_file_path + '/' + 'Hydrogen_PipelineCompressorPowerUsage.tab', format="table", param=model.hydrogenPipelineCompressorElectricityUsage)
    # data.load(filename=tab_file_path + '/' + 'Hydrogen_Distances.tab', format="table", param=model.hydrogenPipelineLength) # Depecreated; Distances are now copied from the transmission distances

    data.load(filename=tab_file_path + '/' + 'Hydrogen_StorageCapitalCost.tab', format="table", param=model.hydrogenStorageCapitalCost)
    data.load(filename=tab_file_path + '/' + 'Hydrogen_StorageFixedOMCost.tab', format="table", param=model.hydrogenStorageFixedOMCost)
    data.load(filename=tab_file_path + '/' + 'Hydrogen_StorageMaxCapacity.tab', format="table", param=model.hydrogenMaxStorageCapacity)

    data.load(filename=tab_file_path + '/' + 'Hydrogen_ElyzerInv.tab', format="table", param=model.elyzerInitInv)
    data.load(filename=tab_file_path + '/' + 'Hydrogen_PipelineInv.tab', format="table", param=model.hydrogenPipeInitInv)
    data.load(filename=tab_file_path + '/' + 'Hydrogen_StorageInv.tab', format="table", param=model.hydrogenStorInitInv)
    

    # data.load(filename=tab_file_path + '/' + 'Hydrogen_Demand.tab', format="table", param=model.hydrogenDemandRaw)

    def prepPipelineLength_rule(model):
        for (n1,n2) in model.HydrogenBidirectionPipelines:
            if (n1,n2) in model.BidirectionalArc:
                model.PipelineLength[n1,n2] = model.transmissionLength[n1,n2]
            elif (n2,n1) in model.BidirectionalArc:
                model.PipelineLength[n1,n2] = model.transmissionLength[n2,n1]
            else:
                print('Error constructing hydrogen pipeline length for bidirectional pipeline ' + n1 + ' and ' + n2)
                exit()
    model.build_PipelineLength = BuildAction(rule=prepPipelineLength_rule)

    def prepElectrolyzerInvCost_rule(model):  #if GREEN_HYDROGEN: Added p in electrolyzerplants atohugh all cost parameters are the same. Consider removing this 
        for i in model.CurrentPeriods:
            costperyear = (model.WACC / (1 - ((1 + model.WACC) ** (1 - model.elyzerLifetime)))) * model.elyzerPlantCapitalCost[i] + model.elyzerFixedOMCost[i] + ((model.WACC/(1-((1+model.WACC)**(1-8)))) + (model.WACC/(1-((1+model.WACC)**(1-16))))) * model.elyzerStackCapitalCost[i]
            costperperiod = costperyear * (1 - (1 + model.discountrate) ** -(min((NoOfPeriods-i+1)*value(model.LeapYearsInvestment), value(model.elyzerLifetime)))) / (1 - (1 / (1 + model.discountrate)))
            model.elyzerInvCost[i] = costperperiod
    model.build_elyzerInvCost = BuildAction(rule=prepElectrolyzerInvCost_rule)

    if REFORMER_HYDROGEN:
        def prepReformerPlantInvCost_rule(model):
            for p in model.ReformerPlants:
                for i in model.CurrentPeriods:
                    costperyear = (model.WACC/(1-((1+model.WACC)**(1-model.ReformerPlantLifetime[p]))))*model.ReformerPlantsCapitalCost[p,i]+model.ReformerPlantFixedOMCost[p,i]
                    costperperiod = costperyear*(1-(1+model.discountrate)**-(min((NoOfPeriods-i+1)*value(model.LeapYearsInvestment), value(model.ReformerPlantLifetime[p]))))/(1-(1/(1+model.discountrate)))
                    model.ReformerPlantInvCost[p,i] = costperperiod
        model.build_ReformerPlantInvCost = BuildAction(rule=prepReformerPlantInvCost_rule)

        def prepReformerMargCost_rule(model):
            for p in model.ReformerPlants:
                for i in model.CurrentPeriods:
                    model.ReformerMargCost[p,i] = model.ReformerPlantVarOMCost[p,i] + model.ReformerFuelCost[p,i]
        model.build_ReformerMargCost = BuildAction(rule=prepReformerMargCost_rule)

        
        def reformer_operational_cost_rule(model, n, i):
            return sum(model.operationalDiscountrate*model.seasScale[s]*model.branchProbab[w]*model.ReformerMargCost[p,i]*model.hydrogenProducedReformer_ton[n,p,i,w,h] for p in model.ReformerPlants for (w,s,h) in model.HoursAndSeasonOfBranch)
        model.reformerOperationalCost = Expression(model.Node, model.CurrentPeriods, rule=reformer_operational_cost_rule)

        def reformer_emissions_rule(model,i,w1,w2,w3,w4,w5,w6,w7,w8,w9,w10,w11,w12): #Calculates tons of CO2 emissions per ton H2 produced with Reformer
            branches = [w1,w2,w3,w4,w5,w6,w7,w8,w9,w10,w11,w12]
            return sum(model.seasScale[s]*model.hydrogenProducedReformer_ton[n,p,i,w,h]*model.ReformerEmissionFactor[p,i] for n in model.Node for p in model.ReformerPlants for (w,s,h) in model.HoursAndSeasonOfBranch if w in branches)
        model.reformerEmissions = Expression(model.CurrentPeriods, model.BranchPath, rule=reformer_emissions_rule)

        def CO2_captured_reformers_rule(model, n, i, w, h):
            return sum(model.ReformerCO2CaptureFactor[r,i] * model.hydrogenProducedReformer_ton[n,r,i,w,h] for r in model.ReformerPlants)
        model.co2_captured_reformers = Expression(model.Node, model.CurrentPeriods, model.HoursOfBranch, rule=CO2_captured_reformers_rule)

    def generators_emissions_rule(model, i, w1,w2,w3,w4,w5,w6,w7,w8,w9,w10,w11,w12):
        branches = [w1,w2,w3,w4,w5,w6,w7,w8,w9,w10,w11,w12]
        return sum(model.seasScale[s]*model.genCO2TypeFactor[g]*(GJperMWh/model.genEfficiency[g,i])*model.genOperational[n,g,i,w,h] for (n,g) in model.GeneratorsOfNode for (w,s,h) in model.HoursAndSeasonOfBranch if w in branches)
    model.generatorEmissions = Expression(model.CurrentPeriods, model.BranchPath, rule=generators_emissions_rule)

    def CO2_captured_generators_rule(model, n, i, w, h):
        return sum(model.genCO2Captured[g] * model.genOperational[n,g,i,w,h] * 3.6 / model.genEfficiency[g,i] for g in model.Generator if (n,g) in model.GeneratorsOfNode)
    model.co2_captured_generators = Expression(model.Node, model.CurrentPeriods, model.HoursOfBranch, rule=CO2_captured_generators_rule)

    def CCS_cost_rule(model,n, i):
        total=0
        total += sum(model.operationalDiscountrate*model.seasScale[s]*model.branchProbab[w]*model.co2_captured_generators[n,i,w,h] for (w,s,h) in model.HoursAndSeasonOfBranch)
        if REFORMER_HYDROGEN:
            total += sum(model.operationalDiscountrate*model.seasScale[s]*model.branchProbab[w]*model.co2_captured_reformers[n,i,w,h] for (w,s,h) in model.HoursAndSeasonOfBranch)
        return total
    model.ccs_cost = Expression(model.Node, model.CurrentPeriods,rule=CCS_cost_rule)

    def prepPipelineInvcost_rule(model):
        for i in model.CurrentPeriods:
            for (n1,n2) in model.HydrogenBidirectionPipelines:
                costperyear= (model.WACC/(1-((1+model.WACC)**(1-model.hydrogenPipelineLifetime))))*model.PipelineLength[n1,n2]*(model.hydrogenPipelineCapCost[i]) + model.PipelineLength[n1,n2]*model.hydrogenPipelineOMCost[i]
                costperperiod =costperyear*(1-(1+model.discountrate)**-(min((NoOfPeriods-i+1)*value(model.LeapYearsInvestment), value(model.hydrogenPipelineLifetime))))/(1-(1/(1+model.discountrate)))
                model.hydrogenPipelineInvCost[n1,n2,i] = costperperiod
    model.build_pipelineInvCost = BuildAction(rule=prepPipelineInvcost_rule)

    def prepHydrogenStorageInvcost_rule(model):
        for i in model.CurrentPeriods:
            costperyear =(model.WACC/(1-((1+model.WACC)**(1-model.hydrogenStorageLifetime))))*model.hydrogenStorageCapitalCost[i]+model.hydrogenStorageFixedOMCost[i]
            costperperiod = costperyear*(1-(1+model.discountrate)**-(min((NoOfPeriods-i+1)*value(model.LeapYearsInvestment), value(model.hydrogenStorageLifetime))))/(1-(1/(1+model.discountrate)))
            model.hydrogenStorageInvCost[i] = costperperiod
    model.build_hydrogenStorageInvCost = BuildAction(rule=prepHydrogenStorageInvcost_rule)

    def prepHydrogenCompressorElectricityUsage_rule(model):
        for (n1,n2) in model.HydrogenBidirectionPipelines:
            model.hydrogenPipelinePowerDemandPerTon[n1,n2] = model.PipelineLength[n1,n2] * model.hydrogenPipelineCompressorElectricityUsage
    model.build_hydrogenPipelineCompressorPowerDemand = BuildAction(rule=prepHydrogenCompressorElectricityUsage_rule)


    ###############
    ##EXPRESSIONS##
    ###############

    def multiplier_rule(model,period):
        coeff=1
        if period>1:
            coeff=pow(1.0+model.discountrate,(-5*(int(period)-1)))
        return coeff
    model.discount_multiplier = Expression(model.CurrentPeriods, rule=multiplier_rule)

    def shed_component_rule(model,n,i):
        return sum(model.operationalDiscountrate*model.seasScale[s]*model.branchProbab[w]*model.nodeLostLoadCost[n,i]*model.loadShed[n,i,w,h] for (w,s,h) in model.HoursAndSeasonOfBranch)
    model.shedcomponent = Expression(model.Node, model.CurrentPeriods, rule=shed_component_rule)

    def operational_cost_rule(model,n,i):
        return sum(model.operationalDiscountrate*model.seasScale[s]*model.branchProbab[w]*model.genMargCost[g,i]*model.genOperational[n,g,i,w,h] for (w,s,h) in model.HoursAndSeasonOfBranch for g in model.Generator if (n,g) in model.GeneratorsOfNode)
    model.operationalcost = Expression(model.Node, model.CurrentPeriods, rule=operational_cost_rule)

    def hydrogen_load_shed_rule(model,n,i):
        hydrogen_curtail_cost = 22000 * hydrogen_MWhPerTon 
        hydrogen_shed_cost = sum(model.seasScale[s] * model.branchProbab[w] * hydrogen_curtail_cost * model.hydrogenDemandShed[n,i,w,h]  for (w,s,h) in model.HoursAndSeasonOfBranch)
        return model.operationalDiscountrate * hydrogen_shed_cost
    model.hydrogen_load_shed_cost = Expression(model.Node, model.CurrentPeriods, rule=hydrogen_load_shed_rule)

    #############
    ##OBJECTIVE##
    #############

    def Obj_rule(model):
        returnSum = sum(model.discount_multiplier[i]*(sum(model.genInvCost[g,i]* model.genInvCap[n,g,i] for (n,g) in model.GeneratorsOfNode) + \
                                                        sum(model.transmissionInvCost[n1,n2,i]*model.transmissionInvCap[n1,n2,i] for (n1,n2) in model.BidirectionalArc) + \
                                                        sum((model.storPWInvCost[b,i]*model.storPWInvCap[n,b,i]+model.storENInvCost[b,i]*model.storENInvCap[n,b,i]) for (n,b) in model.StoragesOfNode) + \
                                                        sum(model.shedcomponent[n,i] for n in model.Node) +\
                                                        sum(model.operationalcost[n,i] for n in model.Node)  + \
                                                        sum(model.elyzerInvCost[i] * model.elyzerCapBuilt[n,i] for n in model.Node) + \
                                                        sum(model.hydrogenPipelineInvCost[n1,n2,i] * model.hydrogenPipelineBuilt[n1,n2,i] for (n1,n2) in model.HydrogenBidirectionPipelines) + \
                                                        sum(model.hydrogenStorageBuilt[n,i] * model.hydrogenStorageInvCost[i] for n in model.Node) + \
                                                        sum(model.hydrogen_load_shed_cost[n,i] for n in model.Node) + \
                                                        sum(model.ccs_cost[n,i]*model.CCSCostTSVariable[i] for n in model.Node)
                                                        )
                        for i in model.CurrentPeriods)
        if REFORMER_HYDROGEN:
            returnSum += sum(model.discount_multiplier[i]*sum((model.ReformerPlantInvCost[p,i] * model.ReformerCapBuilt[n,p,i] + \
                                                               model.reformerOperationalCost[n,i]) for n in model.Node for p in model.ReformerPlants)
                            for i in model.CurrentPeriods)

        return returnSum
    model.Obj = Objective(rule=Obj_rule, sense=minimize)

    ###############
    ##CONSTRAINTS##
    ###############

    def FlowBalance_rule(model, n, i, w, h):
        returnSum = sum(model.genOperational[n, g, i, w, h] for g in model.Generator if (n, g) in model.GeneratorsOfNode) \
                    + sum((model.storageDischargeEff[b] * model.storDischarge[n, b, i, w, h] - model.storCharge[n, b, i, w, h]) for b in model.Storage if (n, b) in model.StoragesOfNode) \
                    + sum((model.lineEfficiency[link, n] * model.transmissionOperational[link, n,  i, w, h] - model.transmissionOperational[n, link, i, w, h]) for link in model.NodesLinked[n]) \
                    - model.sload[n, i, w, h] + model.loadShed[n, i, w, h] \
                    - model.powerForHydrogen[n,i,w,h]
        
        for n2 in model.HydrogenLinks[n]: #Hydrogen pipeline compressor power usage is split 50/50 between sending node and receiving node
            if (n,n2) in model.HydrogenBidirectionPipelines:
                returnSum -= 0.5 * model.hydrogenPipelinePowerDemandPerTon[n,n2] * (model.hydrogenSentPipeline[n,n2,i,w,h] + model.hydrogenSentPipeline[n2,n,i,w,h])
            elif (n2,n) in model.HydrogenBidirectionPipelines:
                returnSum -= 0.5 * model.hydrogenPipelinePowerDemandPerTon[n2,n] * (model.hydrogenSentPipeline[n,n2,i,w,h] + model.hydrogenSentPipeline[n2,n,i,w,h])
        if REFORMER_HYDROGEN and n in model.Node:
            returnSum -= sum(model.ReformerPlantElectricityUse[p,i] * model.hydrogenProducedReformer_ton[n,p,i,w,h] for p in model.ReformerPlants)
        return returnSum == 0
    model.FlowBalance = Constraint(model.Node, model.CurrentPeriods, model.HoursOfBranch, rule=FlowBalance_rule)

    # def max_bio_availability_rule(model, i, w):
    #     bio_use = 0
    #     for n in model.NaturalGasNode:
    #         for g in model.Generator:
    #             if (n,g) in model.GeneratorsOfNode:
    #                 if 'bio' in g.lower():
    #                     if 'cofiring' in g.lower():
    #                         bio_use += sum(model.seasScale[s] * 0.1 * model.genOperational[n,g,i,w,h] / model.genEfficiency[g,i] * GJperMWh for (s,h) in model.HoursOfSeason)
    #                     else:
    #                         bio_use += sum(model.seasScale[s] * model.genOperational[n,g,i,w,h] / model.genEfficiency[g,i] * GJperMWh for (s,h) in model.HoursOfSeason)
    #     return bio_use <= model.availableBioEnergy[i]
    # model.max_bio_availability = Constraint(model.CurrentPeriods, model.Branch, rule=max_bio_availability_rule)

    #################################################################

    def genMaxProd_rule(model, n, g, i, w, h):
        return model.genOperational[n,g,i,w,h] - model.genCapAvail[n,g,w,h,i]*model.genInstalledCap[n,g,i] <= 0
    model.maxGenProduction = Constraint(model.GeneratorsOfNode, model.CurrentPeriods, model.HoursOfBranch, rule=genMaxProd_rule)

    #################################################################

    def ramping_rule(model, n, g, i, w, h):
        if h in model.FirstHoursOfRegSeason:
            return Constraint.Skip
        else:
            if g in model.RampingGenerators:
                return model.genOperational[n,g,i,w,h]-model.genOperational[n,g,i,w,(h-1)] - model.genRampUpCap[g]*model.genInstalledCap[n,g,i] <= 0   #
            else:
                return Constraint.Skip
    model.ramping = Constraint(model.GeneratorsOfNode, model.CurrentPeriods, model.HoursOfBranch, rule=ramping_rule)

    #################################################################
    if SEASONAL_STORAGE:
        def storage_energy_balance_rule(model, n, b, i, w, h):
            if 'hydropumpstorage' in b.lower():
                p = value(model.ParentDictionary[w])
                if h in model.FirstHoursOfRegSeason:
                    if w == p:
                        return Constraint.Skip
                    else:
                        parent_start_level = (model.storOperational[n,b,i,p,h-lengthRegSeason] + model.storDischarge[n,b,i,p,h-lengthRegSeason] - model.storageChargeEff[b]*model.storCharge[n,b,i,p,h-lengthRegSeason])
                        parent_end_level = (model.storageBleedEff[b]*model.storOperational[n,b,i,p,h-1])
                        current_start_level = (model.storOperational[n,b,i,w,h] - model.storageChargeEff[b]*model.storCharge[n,b,i,w,h] + model.storDischarge[n,b,i,w,h])
                        return parent_start_level + (parent_end_level - parent_start_level)*model.seasScale['january'] - current_start_level  == 0   #
                else:
                    return model.storageBleedEff[b]*model.storOperational[n,b,i,w,(h-1)] + model.storageChargeEff[b]*model.storCharge[n,b,i,w,h]-model.storDischarge[n,b,i,w,h]-model.storOperational[n,b,i,w,h] == 0   #
            else:
                if h in model.FirstHoursOfRegSeason:
                    return model.storOperational[n,b,i,w,h]-model.storageChargeEff[b]*model.storCharge[n,b,i,w,h]+model.storDischarge[n,b,i,w,h] - 0.5*model.storENInstalledCap[n,b,i] == 0   #
                else:
                    return model.storageBleedEff[b]*model.storOperational[n,b,i,w,(h-1)] + model.storageChargeEff[b]*model.storCharge[n,b,i,w,h]-model.storDischarge[n,b,i,w,h]-model.storOperational[n,b,i,w,h] == 0   #
        model.storage_energy_balance = Constraint(model.StoragesOfNode, model.CurrentPeriods, model.HoursOfBranch, rule=storage_energy_balance_rule)

        def last_hour_stor_energy_rule(model, n, b, i, w, h):
            if 'hydropumpstorage' in b.lower():
                return Constraint.Skip
            else:
                if (h+1-lengthRegSeason) in model.FirstHoursOfRegSeason:
                    return model.storageBleedEff[b]*model.storOperational[n,b,i,w,h] - 0.5*model.storENInstalledCap[n,b,i] == 0
                else:
                    return Constraint.Skip
        model.last_hour_stor_energy = Constraint(model.StoragesOfNode, model.CurrentPeriods, model.HoursOfBranch, rule=last_hour_stor_energy_rule)
        
        def non_negative_initial_energy_storage_rule(model,n,b,i,w,h):
                if h in model.FirstHoursOfRegSeason:
                    return (model.storOperational[n,b,i,w,h] - model.storageChargeEff[b]*model.storCharge[n,b,i,w,h] + model.storDischarge[n,b,i,w,h]) >= 0
                else:
                    return Constraint.Skip
        model.non_negative_initial_energy_storage = Constraint(model.StoragesOfNode, model.CurrentPeriods, model.HoursOfBranch, rule=non_negative_initial_energy_storage_rule)

        def storage_seasonal_net_zero_balance_rule_hydropump(model, n, b,  i, w1,w2,w3,w4,w5,w6,w7,w8,w9,w10,w11,w12):
            h = max(model.Operationalhour)
            if 'hydropumpstorage' in b.lower():
                parent_start_level = (model.storOperational[n,b,i,w12,h-lengthRegSeason+1] + model.storDischarge[n,b,i,w12,h-lengthRegSeason+1] - model.storageChargeEff[b]*model.storCharge[n,b,i,w12,h-lengthRegSeason+1])
                parent_end_level = (model.storageBleedEff[b]*model.storOperational[n,b,i,w12,h])
                current_start_level = (model.storOperational[n,b,i,w1,1] - model.storageChargeEff[b]*model.storCharge[n,b,i,w1,1] + model.storDischarge[n,b,i,w1,1])
                return parent_start_level + (parent_end_level - parent_start_level)*model.seasScale['january'] - current_start_level  == 0
            else:
                return Constraint.Skip
        model.storage_seasonal_net_zero_balance_hydropump = Constraint(model.StoragesOfNode, model.CurrentPeriods, model.BranchPath, rule=storage_seasonal_net_zero_balance_rule_hydropump)
    else:
        def storage_energy_balance_rule(model, n, b, i, w, h):
            if h in model.FirstHoursOfRegSeason:
                return  model.storOperational[n,b,i,w,h]-model.storageChargeEff[b]*model.storCharge[n,b,i,w,h]+model.storDischarge[n,b,i,w,h] - 0.5*model.storENInstalledCap[n,b,i] == 0   #
            else:
                return model.storageBleedEff[b]*model.storOperational[n,b,i,w,(h-1)] + model.storageChargeEff[b]*model.storCharge[n,b,i,w,h]-model.storDischarge[n,b,i,w,h]-model.storOperational[n,b,i,w,h] == 0   #
        model.storage_energy_balance = Constraint(model.StoragesOfNode, model.CurrentPeriods, model.HoursOfBranch, rule=storage_energy_balance_rule)

        def last_hour_stor_energy_rule(model, n, b, i, w, h):
            if (h+1-lengthRegSeason) in model.FirstHoursOfRegSeason:
                return model.storageBleedEff[b]*model.storOperational[n,b,i,w,h] - 0.5*model.storENInstalledCap[n,b,i] == 0
            else:
                return Constraint.Skip
        model.last_hour_stor_energy = Constraint(model.StoragesOfNode, model.CurrentPeriods, model.HoursOfBranch, rule=last_hour_stor_energy_rule)

    #################################################################

    def storage_operational_cap_rule(model, n, b, i, w, h):
        return model.storOperational[n,b,i,w,h] - model.storENInstalledCap[n,b,i] <= 0   #
    model.storage_operational_cap = Constraint(model.StoragesOfNode, model.CurrentPeriods, model.HoursOfBranch, rule=storage_operational_cap_rule)

    #################################################################

    def storage_power_discharg_cap_rule(model, n, b, i, w, h):
        return model.storDischarge[n,b,i,w,h] - model.storageDiscToCharRatio[b]*model.storPWInstalledCap[n,b,i] <= 0   #
    model.storage_power_discharg_cap = Constraint(model.StoragesOfNode, model.CurrentPeriods, model.HoursOfBranch, rule=storage_power_discharg_cap_rule)

    #################################################################

    def storage_power_charg_cap_rule(model, n, b, i, w, h):
        return model.storCharge[n,b,i,w,h] - model.storPWInstalledCap[n,b,i] <= 0   #
    model.storage_power_charg_cap = Constraint(model.StoragesOfNode, model.CurrentPeriods, model.HoursOfBranch, rule=storage_power_charg_cap_rule)

    #################################################################

    def hydro_gen_limit_rule(model, n, g, i, s, w):
        if g in model.RegHydroGenerator:
            return sum(model.genOperational[n,g,i,w,h] for h in model.Operationalhour if (s,h) in model.HoursOfSeason) - model.maxRegHydroGen[n,i,s,w] <= 0
        else:
            return Constraint.Skip  #
    model.hydro_gen_limit = Constraint(model.GeneratorsOfNode, model.CurrentPeriods, model.BranchesOfSeason, rule=hydro_gen_limit_rule)

    #################################################################

    def hydro_node_limit_rule(model, n, i):
        return sum(model.genOperational[n,g,i,w,h]*model.seasScale[s]*model.branchProbab[w] for g in model.HydroGenerator if (n,g) in model.GeneratorsOfNode for (w,s,h) in model.HoursAndSeasonOfBranch) /1e3 - model.maxHydroNode[n] / 1e3 <= 0   #
    model.hydro_node_limit = Constraint(model.Node, model.CurrentPeriods, rule=hydro_node_limit_rule)


    #################################################################

    def transmission_cap_rule(model, n1, n2, i, w, h):
        if (n1,n2) in model.BidirectionalArc:
            return model.transmissionOperational[(n1,n2),i,w,h] - model.transmissionInstalledCap[(n1,n2),i] <= 0
        elif (n2,n1) in model.BidirectionalArc:
            return model.transmissionOperational[(n1,n2),i,w,h] - model.transmissionInstalledCap[(n2,n1),i] <= 0
    model.transmission_cap = Constraint(model.DirectionalLink, model.CurrentPeriods, model.HoursOfBranch, rule=transmission_cap_rule)

    #################################################################

    if EMISSION_CAP:
        def emission_cap_rule(model, i, w1,w2,w3,w4,w5,w6,w7,w8,w9,w10,w11,w12):
            if REFORMER_HYDROGEN:
                return ((model.generatorEmissions[i,w1,w2,w3,w4,w5,w6,w7,w8,w9,w10,w11,w12] + model.reformerEmissions[i,w1,w2,w3,w4,w5,w6,w7,w8,w9,w10,w11,w12]) / co2_scale_factor) <= model.CO2cap[i] * 1e6 / co2_scale_factor
            else:
                return ((model.generatorEmissions[i,w1,w2,w3,w4,w5,w6,w7,w8,w9,w10,w11,w12]) / co2_scale_factor) <= model.CO2cap[i] * 1e6 / co2_scale_factor
        model.emission_cap = Constraint(model.CurrentPeriods, model.BranchPath, rule=emission_cap_rule)

    #################################################################

    def lifetime_rule_gen(model, n, g, i):
        startPeriod=1
        if value(1+i-(model.genLifetime[g]/model.LeapYearsInvestment))>startPeriod:
            startPeriod=value(1+i-model.genLifetime[g]/model.LeapYearsInvestment)
        return sum(model.genInvCap[n,g,j] for j in model.CurrentPeriods if j>=startPeriod and j<=i) + sum(model.genInitInv[n,g,j] for j in model.Period if j>=startPeriod and j<=i) - model.genInstalledCap[n,g,i] + model.genInitCap[n,g,i]== 0   #
    model.installedCapDefinitionGen = Constraint(model.GeneratorsOfNode, model.CurrentPeriods, rule=lifetime_rule_gen)

    #################################################################

    def lifetime_rule_storEN(model, n, b, i):
        startPeriod=1
        if value(1+i-model.storageLifetime[b]*(1/model.LeapYearsInvestment))>startPeriod:
            startPeriod=value(1+i-model.storageLifetime[b]/model.LeapYearsInvestment)
        return (sum(model.storENInvCap[n,b,j] for j in model.CurrentPeriods if j>=startPeriod and j<=i) + sum(model.storENInitInv[n,b,j] for j in model.Period if j>=startPeriod and j<=i) - model.storENInstalledCap[n,b,i] + model.storENInitCap[n,b,i]) / 1e3 == 0   #
    model.installedCapDefinitionStorEN = Constraint(model.StoragesOfNode, model.CurrentPeriods, rule=lifetime_rule_storEN)

    #################################################################

    def lifetime_rule_storPOW(model, n, b, i):
        startPeriod=1
        if value(1+i-model.storageLifetime[b]*(1/model.LeapYearsInvestment))>startPeriod:
            startPeriod=value(1+i-model.storageLifetime[b]/model.LeapYearsInvestment)
        return sum(model.storPWInvCap[n,b,j] for j in model.CurrentPeriods if j>=startPeriod and j<=i) + sum(model.storPWInitInv[n,b,j] for j in model.Period if j>=startPeriod and j<=i) - model.storPWInstalledCap[n,b,i] + model.storPWInitCap[n,b,i]== 0   #
    model.installedCapDefinitionStorPOW = Constraint(model.StoragesOfNode, model.CurrentPeriods, rule=lifetime_rule_storPOW)

    #################################################################

    def lifetime_rule_trans(model, n1, n2, i):
        startPeriod=1
        if value(1+i-model.transmissionLifetime[n1,n2]*(1/model.LeapYearsInvestment))>startPeriod:
            startPeriod=value(1+i-model.transmissionLifetime[n1,n2]/model.LeapYearsInvestment)
        return sum(model.transmissionInvCap[n1,n2,j] for j in model.CurrentPeriods if j>=startPeriod and j<=i) + sum(model.transInitInv[n1,n2,j] for j in model.Period if j>=startPeriod and j<=i) - model.transmissionInstalledCap[n1,n2,i] + model.transmissionInitCap[n1,n2,i] == 0   #
    model.installedCapDefinitionTrans = Constraint(model.BidirectionalArc, model.CurrentPeriods, rule=lifetime_rule_trans)

    #################################################################

    def investment_gen_cap_rule(model, t, n, i):
        # if value(model.genMaxBuiltCap[n,t,i]) < 2*1e5:
        return sum(model.genInvCap[n,g,i] for g in model.Generator if (n,g) in model.GeneratorsOfNode and (t,g) in model.GeneratorsOfTechnology) - model.genMaxBuiltCap[n,t,i] <= 0
        # else:
        #     return Constraint.Skip
    model.investment_gen_cap = Constraint(model.Technology, model.Node, model.CurrentPeriods, rule=investment_gen_cap_rule)

    #################################################################

    def investment_trans_cap_rule(model, n1, n2, i):
        return model.transmissionInvCap[n1,n2,i] - model.transmissionMaxBuiltCap[n1,n2,i] <= 0
    model.investment_trans_cap = Constraint(model.BidirectionalArc, model.CurrentPeriods, rule=investment_trans_cap_rule)

    #################################################################

    def investment_storage_power_cap_rule(model, n, b, i):
        return model.storPWInvCap[n,b,i] - model.storPWMaxBuiltCap[n,b,i] <= 0
    model.investment_storage_power_cap = Constraint(model.StoragesOfNode, model.CurrentPeriods, rule=investment_storage_power_cap_rule)

    #################################################################

    def investment_storage_energy_cap_rule(model, n, b, i):
        return model.storENInvCap[n,b,i] - model.storENMaxBuiltCap[n,b,i] <= 0
    model.investment_storage_energy_cap = Constraint(model.StoragesOfNode, model.CurrentPeriods, rule=investment_storage_energy_cap_rule)

    ################################################################

    def installed_gen_cap_rule(model, t, n, i):
        # if value(model.genMaxInstalledCap[n,t,i]) < 2*1e5:
        return sum(model.genInstalledCap[n,g,i] for g in model.Generator if (n,g) in model.GeneratorsOfNode and (t,g) in model.GeneratorsOfTechnology) - model.genMaxInstalledCap[n,t,i] <= 0
        # else:
        #     return Constraint.Skip
    model.installed_gen_cap = Constraint(model.Technology, model.Node, model.CurrentPeriods, rule=installed_gen_cap_rule)

    #################################################################

    def installed_trans_cap_rule(model, n1, n2, i):
        return model.transmissionInstalledCap[n1, n2, i] - model.transmissionMaxInstalledCap[n1, n2, i] <= 0
    model.installed_trans_cap = Constraint(model.BidirectionalArc, model.CurrentPeriods, rule=installed_trans_cap_rule)

    #################################################################

    def installed_storage_power_cap_rule(model, n, b, i):
        # if value(model.storPWMaxInstalledCap[n,b,i]) < 1e5:
        return model.storPWInstalledCap[n,b,i] - model.storPWMaxInstalledCap[n,b,i] <= 0
        # else:
        #     return Constraint.Skip
    model.installed_storage_power_cap = Constraint(model.StoragesOfNode, model.CurrentPeriods, rule=installed_storage_power_cap_rule)

    #################################################################

    def installed_storage_energy_cap_rule(model, n, b, i):
        # if value(model.storENMaxInstalledCap[n,b,i]) <= 1.7e6:
        return model.storENInstalledCap[n,b,i] /1e3 - model.storENMaxInstalledCap[n,b,i]/1e3 <= 0
        # else:
        #     return Constraint.Skip
    model.installed_storage_energy_cap = Constraint(model.StoragesOfNode, model.CurrentPeriods, rule=installed_storage_energy_cap_rule)

    #################################################################

    def power_energy_relate_rule(model, n, b, i):
        if b in model.DependentStorage:
            return model.storPWInstalledCap[n,b,i] - model.storagePowToEnergy[b]*model.storENInstalledCap[n,b,i] == 0   #
        else:
            return Constraint.Skip
    model.power_energy_relate = Constraint(model.StoragesOfNode, model.CurrentPeriods, rule=power_energy_relate_rule)

    def shed_limit_rule(model,n,h,i,w):
        return model.loadShed[n,i,w,h] <= model.sload[n,i,w,h]
    # model.shed_limit = Constraint(model.Node, model.CurrentPeriods, model.HoursOfBranch, rule=shed_limit_rule)

    def shedTR_limit_rule(model,n,h,i,w):
        return model.loadShedTR[n,i,w,h] <= model.sloadTR[n,i,w,h]
    # model.shed_limitTR = Constraint(model.Node, model.CurrentPeriods, model.HoursOfBranch, rule=shedTR_limit_rule)

    #################################################################

    # To any reader of this code, the next two constraints are very ugly, and there is likely a better implementation that achieves the same. They were put together as quick fixes, and will be fixed if I remember, have time and can be bothered (in that order of priority). The last is most likely to fail.
    def powerFromHydrogenRule(model, n, g, i, w, h):
        if g in model.HydrogenGenerators:
            return model.genOperational[n,g,i,w,h] == model.genEfficiency[g,i] * model.hydrogenForPower[g,n,i,w,h] * model.hydrogenLHV_ton
        else:
            return Constraint.Skip
    model.powerFromHydrogen = Constraint(model.GeneratorsOfNode, model.CurrentPeriods, model.HoursOfBranch, rule=powerFromHydrogenRule)

    def lifetime_rule_pipeline(model,n1,n2,i):
        startPeriod = 1
        if value(1+i-model.hydrogenPipelineLifetime/model.LeapYearsInvestment)>startPeriod:
            startPeriod=value(1+i-model.hydrogenPipelineLifetime/model.LeapYearsInvestment)
        return sum(model.hydrogenPipelineBuilt[n1,n2,j] for j in model.CurrentPeriods if j>=startPeriod and j<=i) + sum(model.hydrogenPipeInitInv[n1,n2,j] for j in model.Period if j>=startPeriod and j<=i) - model.totalHydrogenPipelineCapacity[n1,n2,i] == 0
    model.installedCapDefinitionPipe = Constraint(model.HydrogenBidirectionPipelines, model.CurrentPeriods, rule=lifetime_rule_pipeline)

    def lifetime_rule_elyzer(model,n,i):
        startPeriod = 1
        if value(1+i-model.elyzerLifetime/model.LeapYearsInvestment)>startPeriod:
            startPeriod=value(1+i-model.elyzerLifetime/model.LeapYearsInvestment)
        return sum(model.elyzerCapBuilt[n,j] for j in model.CurrentPeriods if j>=startPeriod and j<=i) + sum(model.elyzerInitInv[n,j] for j in model.Period if j>=startPeriod and j<=i) - model.elyzerTotalCap[n,i] == 0
    model.installedCapDefinitionElyzer = Constraint(model.Node, model.CurrentPeriods, rule=lifetime_rule_elyzer)

    if REFORMER_HYDROGEN:
        def lifetime_rule_reformer(model,n,p,i):
            startPeriod = 1
            if value(1+i-model.ReformerPlantLifetime[p]/model.LeapYearsInvestment)>startPeriod:
                startPeriod = value(1+i-model.ReformerPlantLifetime[p]/model.LeapYearsInvestment)
            return sum(model.ReformerCapBuilt[n,p,j] for j in model.CurrentPeriods if j>=startPeriod and j<=i) + sum(model.reformerInitInv[n,p,j] for j in model.Period if j>=startPeriod and j<=i) - model.ReformerTotalCap[n,p,i] == 0
        model.installedCapDefinitionReformer = Constraint(model.Node, model.ReformerPlants, model.CurrentPeriods, rule=lifetime_rule_reformer)

    def pipeline_cap_rule(model,n1,n2,i,w,h):
        if (n1,n2) in model.HydrogenBidirectionPipelines:
            return model.hydrogenSentPipeline[(n1,n2),i,w,h] + model.hydrogenSentPipeline[(n2,n1),i,w,h] - model.totalHydrogenPipelineCapacity[(n1,n2),i] <= 0
        elif (n2,n1) in model.HydrogenBidirectionPipelines:
            return model.hydrogenSentPipeline[(n1,n2),i,w,h] + model.hydrogenSentPipeline[(n2,n1),i,w,h] - model.totalHydrogenPipelineCapacity[(n2,n1),i] <= 0
        else:
            print('Problem creating max pipeline capacity constraint for nodes ' + n1 +' and ' + n2)
            exit()
    model.pipeline_cap = Constraint(model.AllowedHydrogenLinks, model.CurrentPeriods, model.HoursOfBranch, rule=pipeline_cap_rule)

    def hydrogen_flow_balance_rule(model,n,i,w,h):
        balance = 0
        balance += sum(model.hydrogenSentPipeline[(n2,n),i,w,h] - model.hydrogenSentPipeline[(n,n2),i,w,h] for n2 in model.HydrogenLinks[n])
        balance -= sum(model.hydrogenForPower[g,n,i,w,h] for g in model.HydrogenGenerators)
        balance += model.hydrogenDischargeStorage[n,i,w,h] - model.hydrogenChargeStorage[n,i,w,h]
        balance += model.hydrogenProducedElectro_ton[n,i,w,h]
        balance -= model.hydrogenDemandMet[n,i,w,h]
        if REFORMER_HYDROGEN:
            balance += sum(model.hydrogenProducedReformer_ton[n,p,i,w,h] for p in model.ReformerPlants)
        return balance == 0
    model.hydrogen_flow_balance = Constraint(model.Node, model.CurrentPeriods, model.HoursOfBranch, rule=hydrogen_flow_balance_rule)

    if FLEX_HYDROGEN:
        def meet_hydrogen_demand_rule(model,n,i,w1,w2,w3,w4,w5,w6,w7,w8,w9,w10,w11,w12):
            branches=[w1,w2,w3,w4,w5,w6,w7,w8,w9,w10,w11,w12]
            return sum(model.seasScale[s] * (model.hydrogenDemandMet[n,i,w,h] + model.hydrogenDemandShed[n,i,w,h]) for (w,s,h) in model.HoursAndSeasonOfBranch if w in branches) == model.h2loadAnnualDemand[n,i] * H2LoadScale
        model.meet_hydrogen_demand = Constraint(model.Node, model.CurrentPeriods, model.BranchPath, rule=meet_hydrogen_demand_rule)
    else:    
        def meet_hydrogen_demand_rule(model,n,i,w,h):
            return model.hydrogenDemandMet[n,i,w,h] + model.hydrogenDemandShed[n,i,w,h] == model.h2load[n,i,w,h]
        model.meet_hydrogen_demand = Constraint(model.Node, model.CurrentPeriods, model.HoursOfBranch, rule=meet_hydrogen_demand_rule)

    def hydrogen_production_rule(model,n,i,w,h):
        return model.hydrogenProducedElectro_ton[n,i,w,h] == model.powerForHydrogen[n,i,w,h] / model.elyzerPowerConsumptionPerTon[i]
    model.hydrogen_production = Constraint(model.Node,model.CurrentPeriods, model.HoursOfBranch, rule=hydrogen_production_rule)

    def hydrogen_production_electrolyzer_capacity_rule(model,n,i,w,h):
        return model.powerForHydrogen[n,i,w,h] <= model.elyzerTotalCap[n,i]
    model.hydrogen_production_electrolyzer_capacity = Constraint(model.Node, model.CurrentPeriods, model.HoursOfBranch, rule=hydrogen_production_electrolyzer_capacity_rule)

    if GREEN_HYDROGEN:
        def lifetime_rule_RESgen(model, n, g, i):
            if g in model.RESGenerators:
                startPeriod=1
                if value(1+i-(model.genLifetime[g]/model.LeapYearsInvestment))>startPeriod:
                    startPeriod=value(1+i-model.genLifetime[g]/model.LeapYearsInvestment)
                return sum(model.genInvCap[n,g,j] for j in model.CurrentPeriods if j>=startPeriod and j<=i) + sum(model.RESGenInitInv[n,g,j] for j in model.Period if j>=startPeriod and j<=i) - model.RESgenInstalledCap[n,g,i] == 0   #
            else:
                return Constraint.Skip
        model.installedRESCapDefinitionGen = Constraint(model.GeneratorsOfNode, model.CurrentPeriods, rule=lifetime_rule_RESgen)

        if RENEWABLE_GRID_RULE:
        # Temporal and spatial rule, remember to add powerForHydrogen in el flow balance!!!
            def maintain_renewable_grid_rule(model,n,i,w1,w2,w3,w4,w5,w6,w7,w8,w9,w10,w11,w12):
                branches = [w1,w2,w3,w4,w5,w6,w7,w8,w9,w10,w11,w12]
                return (0.9 * sum(model.genOperational[n,g,i,w,h] for (w,s,h) in model.HoursAndSeasonOfBranch if w in branches for g in model.Generator if (n,g) in model.GeneratorsOfNode)) <= sum(model.genOperational[n,g,h,i,w] for (w,s,h) in model.HoursAndSeasonOfBranch if w in branches for g in model.RESGenerators if (n,g) in model.GeneratorsOfNode)
            model.maintain_renewable_grid = Constraint(model.RenewableGrid, model.BranchPath, rule=maintain_renewable_grid_rule)

            def hydrogen_additivity_rule(model,n,i):
                return model.elyzerCapBuilt[n,i] <= sum(model.genInstalledCap[n,g,i] for g in model.RESGenerators if (n,g) in model.GeneratorsOfNode)
            model.hydrogen_additivity = Constraint(model.NonRenewableGrid, rule=hydrogen_additivity_rule)
            
            def hydrogen_spatio_temporal_rule(model,n,i,w,h):
                return model.powerForHydrogen[n,i,w,h] <= sum(model.genCapAvail[n,g,w,h,i]*model.RESgenInstalledCap[n,g,i] for g in model.RESGenerators if (n,g) in model.GeneratorsOfNode)
            model.hydrogen_spatio_temporal = Constraint(model.NonRenewableGrid, model.HoursOfBranch, rule=hydrogen_spatio_temporal_rule)
        else:
            def hydrogen_additivity_rule(model,n,i):
                    return model.elyzerCapBuilt[n,i] <= sum(model.genInvCap[n,g,i] for g in model.RESGenerators if (n,g) in model.GeneratorsOfNode)
            model.hydrogen_additivity = Constraint(model.Node, model.CurrentPeriods, rule=hydrogen_additivity_rule)
            
            def hydrogen_spatio_temporal_rule(model,n,i,w,h):
                return model.powerForHydrogen[n,i,w,h] <= sum(model.genCapAvail[n,g,w,h,i]*model.RESgenInstalledCap[n,g,i] for g in model.RESGenerators if (n,g) in model.GeneratorsOfNode)
            model.hydrogen_spatio_temporal = Constraint(model.Node, model.CurrentPeriods, model.HoursOfBranch, rule=hydrogen_spatio_temporal_rule)

    if REFORMER_HYDROGEN:
        def hydrogen_production_reformer_capacity_rule(model,n,p,i,w,h):
            return model.hydrogenProducedReformer_MWh[n,p,i,w,h] <= model.ReformerTotalCap[n,p,i]
        model.hydrogen_production_reformer_capacity = Constraint(model.Node, model.ReformerPlants, model.CurrentPeriods, model.HoursOfBranch, rule=hydrogen_production_reformer_capacity_rule)

        def hydrogen_link_reformer_ton_MWh_rule(model,n,p,i,w,h):
            return model.hydrogenProducedReformer_ton[n,p,i,w,h] == model.hydrogenProducedReformer_MWh[n,p,i,w,h]/model.hydrogenLHV_ton
        model.hydrogen_link_reformer_ton_MWh = Constraint(model.Node, model.ReformerPlants, model.CurrentPeriods, model.HoursOfBranch, rule=hydrogen_link_reformer_ton_MWh_rule)

        def hydrogen_reformer_ramp_rule(model,n,p,i,w,h):
            if h in model.FirstHoursOfRegSeason:
                return Constraint.Skip
            else:
                return model.hydrogenProducedReformer_MWh[n,p,i,w,h] - model.hydrogenProducedReformer_MWh[n,p,i,w,h-1] <= 0.1 * model.ReformerTotalCap[n,p,i]
        model.hydrogen_reformer_ramp = Constraint(model.Node, model.ReformerPlants, model.CurrentPeriods, model.HoursOfBranch, rule=hydrogen_reformer_ramp_rule)

    if SEASONAL_STORAGE:
        def hydrogen_storage_balance_rule(model,n,i,w,h):
            p = value(model.ParentDictionary[w])
            if h in model.FirstHoursOfRegSeason:
                if w == p:
                    #return model.hydrogenStorageInitOperational * model.hydrogenTotalStorage[n,i]+model.hydrogenChargeStorage[n,i,w,h]-model.hydrogenDischargeStorage[n,i,w,h] - model.hydrogenStorageOperational[n,i,w,h] == 0
                    return Constraint.Skip
                else:
                    parent_start_level = (model.hydrogenStorageOperational[n,i,p,h-lengthRegSeason]+model.hydrogenDischargeStorage[n,i,p,h-lengthRegSeason]-model.hydrogenChargeStorage[n,i,p,h-lengthRegSeason])
                    parent_end_level = (model.hydrogenStorageOperational[n,i,p,h-1])
                    current_start_level = (model.hydrogenStorageOperational[n,i,w,h] - model.hydrogenChargeStorage[n,i,w,h] + model.hydrogenDischargeStorage[n,i,w,h])
                    return parent_start_level + (parent_end_level - parent_start_level)*model.seasScale['january'] - current_start_level == 0
            else:
                return model.hydrogenStorageOperational[n,i,w,h-1] + model.hydrogenChargeStorage[n,i,w,h] - model.hydrogenDischargeStorage[n,i,w,h] - model.hydrogenStorageOperational[n,i,w,h] == 0
        model.hydrogen_storage_balance = Constraint(model.Node, model.CurrentPeriods, model.HoursOfBranch, rule=hydrogen_storage_balance_rule)

        def hydrogen_balance_storage_rule(model,n,i,w1,w2,w3,w4,w5,w6,w7,w8,w9,w10,w11,w12):
            h = max(model.Operationalhour)
            parent_start_level = (model.hydrogenStorageOperational[n,i,w12,h-lengthRegSeason+1]+model.hydrogenDischargeStorage[n,i,w12,h-lengthRegSeason+1]-model.hydrogenChargeStorage[n,i,w12,h-lengthRegSeason+1])
            parent_end_level = (model.hydrogenStorageOperational[n,i,w12,h])
            current_start_level = (model.hydrogenStorageOperational[n,i,w1,1] - model.hydrogenChargeStorage[n,i,w1,1] + model.hydrogenDischargeStorage[n,i,w1,1])
            return parent_start_level + (parent_end_level - parent_start_level)*model.seasScale['january'] - current_start_level == 0
        model.hydrogen_balance_storage = Constraint(model.Node, model.CurrentPeriods, model.BranchPath, rule=hydrogen_balance_storage_rule)

        def non_negative_initial_hydrogen_storage_rule(model,n,i,w,h):
            if h in model.FirstHoursOfRegSeason:
                return (model.hydrogenStorageOperational[n,i,w,h] - model.hydrogenChargeStorage[n,i,w,h] + model.hydrogenDischargeStorage[n,i,w,h]) >= 0
            else:
                return Constraint.Skip
        model.non_negative_initial_hydrogen_storage = Constraint(model.Node, model.CurrentPeriods, model.HoursOfBranch, rule=non_negative_initial_hydrogen_storage_rule)
    else:
        def hydrogen_storage_balance_rule(model,n,i,w,h):
            if h in model.FirstHoursOfRegSeason:
                return  model.hydrogenStorageOperational[n,i,w,h] - model.hydrogenChargeStorage[n,i,w,h] + model.hydrogenDischargeStorage[n,i,w,h] - 0.5*model.hydrogenTotalStorage[n,i] == 0
            else:
                return model.hydrogenStorageOperational[n,i,w,h-1] + model.hydrogenChargeStorage[n,i,w,h] - model.hydrogenDischargeStorage[n,i,w,h] - model.hydrogenStorageOperational[n,i,w,h] == 0
        model.hydrogen_storage_balance = Constraint(model.Node, model.CurrentPeriods, model.HoursOfBranch, rule=hydrogen_storage_balance_rule)

        def last_hour_hydrogen_storage_rule(model,n,i,w,h):
            if (h+1-lengthRegSeason) in model.FirstHoursOfRegSeason:
                return model.hydrogenStorageOperational[n,i,w,h] - 0.5*model.hydrogenTotalStorage[n,i] == 0
            else:
                return Constraint.Skip
        model.last_hour_hydrogen_storage = Constraint(model.Node, model.CurrentPeriods, model.HoursOfBranch, rule=last_hour_hydrogen_storage_rule)

    def hydrogen_storage_charge_rule(model,n,i,w,h):
        return (model.hydrogenChargeStorage[n,i,w,h] + model.hydrogenDischargeStorage[n,i,w,h])/1e3 <= MaxChargeAndDischargePercentage*model.hydrogenTotalStorage[n,i] /1e3
    model.hydrogen_storage_charge = Constraint(model.Node, model.CurrentPeriods, model.HoursOfBranch, rule=hydrogen_storage_charge_rule)

    def hydrogen_storage_max_capacity_rule(model,n,i):
        return model.hydrogenTotalStorage[n,i] /1e3 <= model.hydrogenMaxStorageCapacity[n] /1e3
    model.hydrogen_storage_max_capacity = Constraint(model.Node, model.CurrentPeriods, rule=hydrogen_storage_max_capacity_rule)

    def hydrogen_storage_operational_capacity_rule(model,n,i,w,h):
        return model.hydrogenStorageOperational[n,i,w,h] <= model.hydrogenTotalStorage[n,i]
    model.hydrogen_storage_operational_capacity = Constraint(model.Node, model.CurrentPeriods, model.HoursOfBranch, rule=hydrogen_storage_operational_capacity_rule)

    def hydrogen_storage_lifetime_rule(model,n,i):
        startPeriod=1
        if value(1+i-model.hydrogenStorageLifetime/model.LeapYearsInvestment)>startPeriod:
            startPeriod = value(1+i-model.hydrogenStorageLifetime/model.LeapYearsInvestment)
        return sum(model.hydrogenStorageBuilt[n,j] for j in model.CurrentPeriods if j>=startPeriod and j <=i) + sum(model.hydrogenStorInitInv[n,j] for j in model.Period if j>=startPeriod and j <=i) - model.hydrogenTotalStorage[n,i] == 0
    model.hydrogen_storage_lifetime = Constraint(model.Node, model.CurrentPeriods, rule=hydrogen_storage_lifetime_rule)

    stopConstraints = startBuild = datetime.now()
    #######
    ##RUN##
    #######

    print("Objective and constraints read...")

    print("{hour}:{minute}:{second}: Building instance...".format(
        hour=datetime.now().strftime("%H"), minute=datetime.now().strftime("%M"), second=datetime.now().strftime("%S")))

    start = time.time()

    instance = model.create_instance(data) #, report_timing=True)
    instance.dual = Suffix(direction=Suffix.IMPORT) #Make sure the dual value is collected into solver results (if solver supplies dual information)

    inv_per = []
    for i in instance.Period:
        my_string = str(value(start_year+int(i)*LeapYearsInvestment))+"-"+str(value(start_year+LeapYearsInvestment+int(i)*LeapYearsInvestment))
        inv_per.append(my_string)

    print("{hour}:{minute}:{second}: Writing load data to data_electric_load.csv...".format(
        hour=datetime.now().strftime("%H"), minute=datetime.now().strftime("%M"), second=datetime.now().strftime("%S")))
    f = open(tab_file_path + "/" + 'data_electric_load.csv', 'w', newline='')
    writer = csv.writer(f)
    my_header = ["Node","Period","Branch","Season","Hour",'Electric load [MW]']
    writer.writerow(my_header)
    for n in instance.Node:
        for i in instance.CurrentPeriods:
                for (w,s,h) in instance.HoursAndSeasonOfBranch:
                    my_string = [n,inv_per[int(i-1)],w,s,h,
                                 value(instance.sload[n,i,w,h])]
                    writer.writerow(my_string)
    f.close()

    end = time.time()
    print("{hour}:{minute}:{second}: Building instance took [sec]:".format(
        hour=datetime.now().strftime("%H"), minute=datetime.now().strftime("%M"), second=datetime.now().strftime("%S")) + str(end - start))

    endBuild = startOptimization = datetime.now()

    #import pdb; pdb.set_trace()

    print("\n----------------------Problem Statistics---------------------")
    print("Nodes: "+ str(len(instance.Node)))
    print("Lines: "+str(len(instance.BidirectionalArc)))
    print("")
    print("GeneratorTypes: "+str(len(instance.Generator)))
    print("TotalGenerators: "+str(len(instance.GeneratorsOfNode)))
    print("StorageTypes: "+str(len(instance.Storage)))
    print("TotalStorages: "+str(len(instance.StoragesOfNode)))
    print("")
    print("InvestmentYears: "+str(len(instance.CurrentPeriods)))
    print("Branchs: "+str(len(instance.Branch)))
    print("TotalOperationalHoursPerBranch: "+str(len(instance.Operationalhour)))
    print("TotalOperationalHoursPerInvYear: "+str(len(instance.Operationalhour)*len(instance.Branch)))
    print("Seasons: "+str(len(instance.Season)))
    print("RegularSeasons: "+str(len(instance.FirstHoursOfRegSeason)))
    print("lengthRegSeason: "+str(value(instance.lengthRegSeason)))
    print("")
    print("Discount rate: "+str(value(instance.discountrate)))
    print(f"Operational discount scale: {value(instance.operationalDiscountrate):.3f}")
    print("Seasonal scale: " + str(value(instance.seasScale['january'])))
    # print("Optimizing with hydrogen: " + str(hydrogen))
    print("--------------------------------------------------------------\n")

    if WRITE_LP:
        print("Writing LP-file...")
        start = time.time()
        lpstring = 'LP_' + name + '.lp'
        if USE_TEMP_DIR:
            lpstring = './LP_'+ name + '.lp'
        instance.write(lpstring, io_options={'symbolic_solver_labels': True})
        end = time.time()
        print("Writing LP-file took [sec]:")
        print(end - start)

    print("{hour}:{minute}:{second}: Solving...".format(
        hour=datetime.now().strftime("%H"), minute=datetime.now().strftime("%M"), second=datetime.now().strftime("%S")))

    if solver == "CPLEX":
        opt = SolverFactory("cplex", Verbose=True)
        opt.options["lpmethod"] = 4
        opt.options["barrier crossover"] = -1
        #instance.display('outputs_cplex.txt')
    if solver == "Xpress":
        opt = SolverFactory("xpress") #Verbose=True
        opt.options["defaultAlg"] = 4
        opt.options["crossover"] = 0
        opt.options["lpLog"] = 1
        opt.options["Trace"] = 1
        #instance.display('outputs_xpress.txt')
    if solver == "Gurobi":
        opt = SolverFactory('gurobi', Verbose=True)
        opt.options["Crossover"]=0
        # opt.options['NumericFocus']=1
        # opt.options['BarHomogeneous']=1
        # opt.options['Presolve']=2
        # opt.options['FeasibilityTol']=10**(-9)
        opt.options["Method"]=2
        opt.options["BarConvTol"]=1e-5
        opt.options['ResultFile'] = f"{name}.ilp"

    results = opt.solve(instance, tee=True, logfile=result_file_path + '/logfile_' + name + '.log')#, keepfiles=True, symbolic_solver_labels=True)

    endOptimization = StartReporting = datetime.now()

    #instance.display('outputs_gurobi.txt')

    #import pdb; pdb.set_trace()

    ######################
    ##UPDATE INVESTMENTS##
    ######################

    print("{hour}:{minute}:{second}: Updating investments to file\n".format(
            hour=datetime.now().strftime("%H"), minute=datetime.now().strftime("%M"), second=datetime.now().strftime("%S")))
    
    #GeneratorInv
    data = []
    for (n, g) in instance.GeneratorsOfNode:
        for i in instance.Period:
            inv_value = value(instance.genInitInv[n, g, i])
            if value(i) in updatePeriods:
                inv_value += value(instance.genInvCap[n, g, i])
            data.append((n, g, i, inv_value))
    
    df = pd.DataFrame(data, columns=['Node', 'Generator', 'Period', 'Capacity'])

    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), data_handler_path, 'Generator.xlsx')
    wb = load_workbook(file_path)
    sheet = wb['GeneratorInv']
    header_row = 3
    for col, header in enumerate(df.columns, start=1):
        sheet.cell(row=header_row, column=col, value=header)
    start_row = 4
    for index, row in df.iterrows():
        sheet.cell(row=start_row + index, column=1, value=row['Node'])
        sheet.cell(row=start_row + index, column=2, value=row['Generator'])
        sheet.cell(row=start_row + index, column=3, value=row['Period'])
        sheet.cell(row=start_row + index, column=4, value=row['Capacity'])
    wb.save(file_path)

    #transmissionInv
    data = []
    for (n1, n2) in instance.BidirectionalArc:
        for i in instance.Period:
            inv_value = value(instance.transInitInv[n1, n2, i])
            if value(i) in updatePeriods:
                inv_value += value(instance.transmissionInvCap[n1,n2,i])
            data.append((n1, n2, i, inv_value))
    
    df = pd.DataFrame(data, columns=['NodeOne', 'NodeTwo', 'Period', 'Capacity'])

    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), data_handler_path, 'Transmission.xlsx')
    wb = load_workbook(file_path)
    sheet = wb['TransmissionInv']
    header_row = 3
    for col, header in enumerate(df.columns, start=1):
        sheet.cell(row=header_row, column=col, value=header)
    start_row = 4
    for index, row in df.iterrows():
        sheet.cell(row=start_row + index, column=1, value=row['NodeOne'])
        sheet.cell(row=start_row + index, column=2, value=row['NodeTwo'])
        sheet.cell(row=start_row + index, column=3, value=row['Period'])
        sheet.cell(row=start_row + index, column=4, value=row['Capacity'])
    wb.save(file_path)

    #StoragePWInv
    data = []
    for (n,b) in instance.StoragesOfNode:
        for i in instance.Period:
            inv_value = value(instance.storPWInitInv[n, b, i])
            if value(i) in updatePeriods:
                inv_value += value(instance.storPWInvCap[n,b,i])
            data.append((n, b, i, inv_value))
    
    df = pd.DataFrame(data, columns=['Node', 'Storage', 'Period', 'Capacity'])

    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), data_handler_path, 'Storage.xlsx')
    wb = load_workbook(file_path)
    sheet = wb['StoragePWInv']
    header_row = 3
    for col, header in enumerate(df.columns, start=1):
        sheet.cell(row=header_row, column=col, value=header)
    start_row = 4
    for index, row in df.iterrows():
        sheet.cell(row=start_row + index, column=1, value=row['Node'])
        sheet.cell(row=start_row + index, column=2, value=row['Storage'])
        sheet.cell(row=start_row + index, column=3, value=row['Period'])
        sheet.cell(row=start_row + index, column=4, value=row['Capacity'])
    wb.save(file_path)

    #StorageENInv
    data = []
    for (n,b) in instance.StoragesOfNode:
        for i in instance.Period:
            inv_value = value(instance.storENInitInv[n, b, i])
            if value(i) in updatePeriods:
                inv_value += value(instance.storENInvCap[n,b,i])
            data.append((n, b, i, inv_value))
    
    df = pd.DataFrame(data, columns=['Node', 'Storage', 'Period', 'Capacity'])

    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), data_handler_path, 'Storage.xlsx')
    wb = load_workbook(file_path)
    sheet = wb['StorageENInv']
    header_row = 3
    for col, header in enumerate(df.columns, start=1):
        sheet.cell(row=header_row, column=col, value=header)
    start_row = 4
    for index, row in df.iterrows():
        sheet.cell(row=start_row + index, column=1, value=row['Node'])
        sheet.cell(row=start_row + index, column=2, value=row['Storage'])
        sheet.cell(row=start_row + index, column=3, value=row['Period'])
        sheet.cell(row=start_row + index, column=4, value=row['Capacity'])
    wb.save(file_path)

    #ElyzerInv
    data = []
    for n in instance.Node:
        for i in instance.Period:
            inv_value = value(instance.elyzerInitInv[n, i])
            if value(i) in updatePeriods:
                inv_value += value(instance.elyzerCapBuilt[n,i])
            data.append((n, i, inv_value))
    
    df = pd.DataFrame(data, columns=['Node', 'Period', 'Capacity'])

    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), data_handler_path, 'Hydrogen.xlsx')
    wb = load_workbook(file_path)
    sheet = wb['ElyzerInv']
    header_row = 3
    for col, header in enumerate(df.columns, start=1):
        sheet.cell(row=header_row, column=col, value=header)
    start_row = 4
    for index, row in df.iterrows():
        sheet.cell(row=start_row + index, column=1, value=row['Node'])
        sheet.cell(row=start_row + index, column=2, value=row['Period'])
        sheet.cell(row=start_row + index, column=3, value=row['Capacity'])
    wb.save(file_path)

    #PipelineInv
    data = []
    for (n1, n2) in instance.HydrogenBidirectionPipelines:
        for i in instance.Period:
            inv_value = value(instance.hydrogenPipeInitInv[n1, n2, i])
            if value(i) in updatePeriods:
                inv_value += value(instance.hydrogenPipelineBuilt[n1,n2,i])
            data.append((n1, n2, i, inv_value))
    
    df = pd.DataFrame(data, columns=['NodeOne', 'NodeTwo', 'Period', 'Capacity'])

    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), data_handler_path, 'Hydrogen.xlsx')
    wb = load_workbook(file_path)
    sheet = wb['PipelineInv']
    header_row = 3
    for col, header in enumerate(df.columns, start=1):
        sheet.cell(row=header_row, column=col, value=header)
    start_row = 4
    for index, row in df.iterrows():
        sheet.cell(row=start_row + index, column=1, value=row['NodeOne'])
        sheet.cell(row=start_row + index, column=2, value=row['NodeTwo'])
        sheet.cell(row=start_row + index, column=3, value=row['Period'])
        sheet.cell(row=start_row + index, column=4, value=row['Capacity'])
    wb.save(file_path)

    #StorageInv
    data = []
    for n in instance.Node:
        for i in instance.Period:
            inv_value = value(instance.hydrogenStorInitInv[n,i])
            if value(i) in updatePeriods:
                inv_value += value(instance.hydrogenStorageBuilt[n,i])
            data.append((n, i, inv_value))
    
    df = pd.DataFrame(data, columns=['Node', 'Period', 'Capacity'])

    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), data_handler_path, 'Hydrogen.xlsx')
    wb = load_workbook(file_path)
    sheet = wb['StorageInv']
    header_row = 3
    for col, header in enumerate(df.columns, start=1):
        sheet.cell(row=header_row, column=col, value=header)
    start_row = 4
    for index, row in df.iterrows():
        sheet.cell(row=start_row + index, column=1, value=row['Node'])
        sheet.cell(row=start_row + index, column=2, value=row['Period'])
        sheet.cell(row=start_row + index, column=3, value=row['Capacity'])
    wb.save(file_path)

    #RESGeneratorInv
    data = []
    for (n, g) in instance.GeneratorsOfNode:
        for i in instance.Period:
            if g in instance.RESGenerators:
                inv_value = value(instance.RESGenInitInv[n, g, i])
                if value(i) in updatePeriods:
                    inv_value += value(instance.genInvCap[n, g, i])
                data.append((n, g, i, inv_value))
    
    df = pd.DataFrame(data, columns=['Node', 'Generator', 'Period', 'Capacity'])

    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), data_handler_path, 'Generator.xlsx')
    wb = load_workbook(file_path)
    sheet = wb['RESGeneratorInv']
    header_row = 3
    for col, header in enumerate(df.columns, start=1):
        sheet.cell(row=header_row, column=col, value=header)
    start_row = 4
    for index, row in df.iterrows():
        sheet.cell(row=start_row + index, column=1, value=row['Node'])
        sheet.cell(row=start_row + index, column=2, value=row['Generator'])
        sheet.cell(row=start_row + index, column=3, value=row['Period'])
        sheet.cell(row=start_row + index, column=4, value=row['Capacity'])
    wb.save(file_path)

    #ReformerInv
    if REFORMER_HYDROGEN:
        data = []
        for n in instance.Node:
            for p in instance.ReformerPlants:
                for i in instance.Period:
                    inv_value = value(instance.reformerInitInv[n, p, i])
                    if value(i) in updatePeriods:
                        inv_value += value(instance.ReformerCapBuilt[n,p,i])
                    data.append((n, p, i, inv_value))
        
        df = pd.DataFrame(data, columns=['Node', 'Plant', 'Period', 'Capacity'])

        file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), data_handler_path, 'Hydrogen.xlsx')
        wb = load_workbook(file_path)
        sheet = wb['ReformerInv']
        header_row = 3
        for col, header in enumerate(df.columns, start=1):
            sheet.cell(row=header_row, column=col, value=header)
        start_row = 4
        for index, row in df.iterrows():
            sheet.cell(row=start_row + index, column=1, value=row['Node'])
            sheet.cell(row=start_row + index, column=2, value=row['Plant'])
            sheet.cell(row=start_row + index, column=3, value=row['Period'])
            sheet.cell(row=start_row + index, column=4, value=row['Capacity'])
        wb.save(file_path)


    ###########
    ##RESULTS##
    ###########

    def calculatePowerEmissionIntensity(n,h,i,w,m=None): # kg CO2 for power consumption
        #print(f'Evaluating {n}')
        emissions = 1000 * value(sum(instance.genOperational[n,g,i,w,h]*instance.genCO2TypeFactor[g]*(GJperMWh/instance.genEfficiency[g,i]) for g in instance.Generator if (n,g) in instance.GeneratorsOfNode))
        total_power = value(sum(instance.genOperational[n,g,i,w,h] for g in instance.Generator if (n,g) in instance.GeneratorsOfNode))
        for n2 in instance.NodesLinked[n]:
            if value(instance.lineEfficiency[n2,n]*instance.transmissionOperational[n2,n,i,w,h]) > 200 and value(instance.lineEfficiency[n,n2]*instance.transmissionOperational[n,n2,i,w,h]) < 1:
                if n2==m:
                    print(f'Warning: We have recursion loop between {n} and {m} in calculatePowerEmissionIntensity!')
                emissions += calculatePowerEmissionIntensity(n2,h,i,w,n) * value(instance.lineEfficiency[n2,n]*instance.transmissionOperational[n2,n,i,w,h])
                total_power += value(instance.lineEfficiency[n2,n]*instance.transmissionOperational[n2,n,i,w,h])
            else:
                emissions += 0
                total_power += 0
        if total_power > 0:
            emission_factor = emissions/total_power
        else:
            emission_factor = 0
            # print(f'Warning: Total power in {n} in hour {h} in {inv_per[int(i-1)]} in {w} is 0!')
        #print(f'Node {n}, period {i}, hour {h}, {w}:\tEm.fac.:{emission_factor:.3f}')
        return emission_factor

    try:
        print(("{hour}:{minute}:{second}: Writing results in " + result_file_path + '/\n').format(
            hour=datetime.now().strftime("%H"), minute = datetime.now().strftime("%M"), second=datetime.now().strftime("%S")))

        if 'results_objective' in include_results:
            print("{hour}:{minute}:{second}: Writing objective functions results in results_objective.csv...".format(
                hour=datetime.now().strftime("%H"), minute=datetime.now().strftime("%M"), second=datetime.now().strftime("%S")))
            f = open(result_file_path + "/" + 'results_objective.csv', 'a', newline='')
            writer = csv.writer(f)
            writer.writerow(["Objective function value:" + str(value(instance.Obj))])
            writer.writerow(["Scientific notation:", str(value(instance.Obj))])
            writer.writerow(["Solver status:",results.solver.status])
            f.close()
        
        if 'results_objective_detailed' in include_results:
            print("{hour}:{minute}:{second}: Writing objective functions results in results_objective_detailed.csv...".format(
                hour=datetime.now().strftime("%H"), minute=datetime.now().strftime("%M"), second=datetime.now().strftime("%S")))
            f = open(result_file_path + "/" + 'results_objective_detailed.csv', 'a', newline='')
            writer = csv.writer(f)
            if instance.CurrentPeriods[1]==1:
                header = ["Node","Period","GeneratorInvCost","PowerStorageInvCost","ShedCost",
                                 "OperationalCost","ElyzerInvCost","HydrogenStorageInvCost",
                                "HydrogenLoadShed","CCScost"]
                if REFORMER_HYDROGEN:
                    header.extend(["ReformerPlantInvCost","ReformerOperationalCost"])
                writer.writerow(header)
            for n in instance.Node:
                for i in instance.CurrentPeriods:   
                    if i in updatePeriods or last_run == True:
                        row = [n,inv_per[int(i-1)],
                        value(instance.discount_multiplier[i]*sum(instance.genInvCost[g,i]* instance.genInvCap[n,g,i] for g in instance.Generator if (n,g) in instance.GeneratorsOfNode)),
                        value(instance.discount_multiplier[i]*sum((instance.storPWInvCost[b,i]*instance.storPWInvCap[n,b,i]+instance.storENInvCost[b,i]*instance.storENInvCap[n,b,i]) for b in instance.Storage if (n,b) in instance.StoragesOfNode)),
                        value(instance.discount_multiplier[i]*instance.shedcomponent[n,i]),
                        value(instance.discount_multiplier[i]*instance.operationalcost[n,i]),
                        value((instance.discount_multiplier[i]*instance.elyzerInvCost[i] * instance.elyzerCapBuilt[n,i])),
                        value(instance.discount_multiplier[i]*instance.hydrogenStorageBuilt[n,i] * instance.hydrogenStorageInvCost[i]),
                        value(instance.discount_multiplier[i]*instance.hydrogen_load_shed_cost[n,i]),
                        value(instance.discount_multiplier[i]*instance.ccs_cost[n,i]*instance.CCSCostTSVariable[i])
                        ]
                        if REFORMER_HYDROGEN:
                            row.extend([value(instance.discount_multiplier[i]*sum(instance.ReformerPlantInvCost[p,i] * instance.ReformerCapBuilt[n,p,i] for p in instance.ReformerPlants)),
                                        value(instance.discount_multiplier[i]*instance.reformerOperationalCost[n,i])])
                        writer.writerow(row)
            f.close()
            
        if 'results_objective_transmission' in include_results:
            print("{hour}:{minute}:{second}: Writing objective functions results in results_objective_transmission.csv...".format(
                hour=datetime.now().strftime("%H"), minute=datetime.now().strftime("%M"), second=datetime.now().strftime("%S")))
            f = open(result_file_path + "/" + 'results_objective_transmission.csv', 'a', newline='')
            writer = csv.writer(f)
            if instance.CurrentPeriods[1]==1:
                writer.writerow(["FromNode","ToNode","Period","TransmissionInvCost","HydrogenPipelineInvCost"])
            for n1 in instance.Node:
                for n2 in instance.Node:
                    for i in instance.CurrentPeriods:   
                        if i in updatePeriods or last_run == True:
                            writer.writerow([n1,n2,inv_per[int(i-1)],
                            value(instance.discount_multiplier[i]*instance.transmissionInvCost[n1,n2,i]*instance.transmissionInvCap[n1,n2,i] if (n1,n2) in instance.BidirectionalArc else 0),
                            value(instance.discount_multiplier[i]*instance.hydrogenPipelineInvCost[n1,n2,i] * instance.hydrogenPipelineBuilt[n1,n2,i] if (n1,n2) in instance.HydrogenBidirectionPipelines else 0),
                            ])
            f.close()

        if 'results_output_transmission' in include_results:
            print("{hour}:{minute}:{second}: Writing transmission investment decisions to results_output_transmission.csv...".format(
                hour=datetime.now().strftime("%H"), minute=datetime.now().strftime("%M"), second=datetime.now().strftime("%S")))
            f = open(result_file_path + "/" + 'results_output_transmission.csv', 'a', newline='')
            writer = csv.writer(f)
            if instance.CurrentPeriods[1]==1:
                writer.writerow(["BetweenNode","AndNode","Period","transmissionInvCap_MW","transmissionInvCapMax_MW","transmissionInstalledCap_MW","transmissionInstalledCapMax_MW","DiscountedInvestmentCost_Euro","transmissionExpectedAnnualVolume_GWh","ExpectedAnnualLosses_GWh"])
            for (n1,n2) in instance.BidirectionalArc:
                for i in instance.CurrentPeriods:   
                    if i in updatePeriods or last_run == True:
                        writer.writerow([n1,n2,inv_per[int(i-1)],
                                            value(instance.transmissionInvCap[n1,n2,i]), value(instance.transmissionMaxBuiltCap[n1,n2,i]),
                                            value(instance.transmissionInstalledCap[n1,n2,i]), value(instance.transmissionMaxInstalledCap[n1,n2,i]),
                                            value(instance.discount_multiplier[i]*instance.transmissionInvCap[n1,n2,i]*instance.transmissionInvCost[n1,n2,i]),
                                            value(sum(instance.branchProbab[w]*instance.seasScale[s]*(instance.transmissionOperational[n1,n2,i,w,h]+instance.transmissionOperational[n2,n1,i,w,h])/1000 for (w,s,h) in instance.HoursAndSeasonOfBranch)),
                                            value(sum(instance.branchProbab[w]*instance.seasScale[s]*((1 - instance.lineEfficiency[n1,n2])*instance.transmissionOperational[n1,n2,i,w,h] + (1 - instance.lineEfficiency[n2,n1])*instance.transmissionOperational[n2,n1,i,w,h])/1000 for (w,s,h) in instance.HoursAndSeasonOfBranch))])
            f.close()

        if 'results_output_gen' in include_results:
            print("{hour}:{minute}:{second}: Writing generator investment decisions to results_output_gen.csv...".format(
                    hour=datetime.now().strftime("%H"), minute=datetime.now().strftime("%M"), second=datetime.now().strftime("%S")))
            f = open(result_file_path + "/" + 'results_output_gen.csv', 'a', newline='')
            writer = csv.writer(f)
            if instance.CurrentPeriods[1]==1:
                my_string = ["Node","GeneratorType","Period","genInvCap_MW","genInstalledCap_MW","genExpectedCapacityFactor","DiscountedInvestmentCost_Euro","genExpectedAnnualProduction_GWh"]
                writer.writerow(my_string)
            for (n,g) in instance.GeneratorsOfNode:
                for i in instance.CurrentPeriods:               
                    if i in updatePeriods or last_run == True:
                        my_string=[n,g,inv_per[int(i-1)],value(instance.genInvCap[n,g,i]),value(instance.genInstalledCap[n,g,i]),
                                    value(sum(instance.branchProbab[w]*instance.seasScale[s]*instance.genOperational[n,g,i,w,h] for (w,s,h) in instance.HoursAndSeasonOfBranch)/(instance.genInstalledCap[n,g,i]*8760) if value(instance.genInstalledCap[n,g,i]) != 0 and value(instance.genInstalledCap[n,g,i]) > 3 else 0),
                                    value(instance.discount_multiplier[i]*instance.genInvCap[n,g,i]*instance.genInvCost[g,i])]
                        my_string.append(value(sum(instance.seasScale[s]*instance.branchProbab[w]*instance.genOperational[n,g,i,w,h]/1000 for (w,s,h) in instance.HoursAndSeasonOfBranch) if value(instance.genInstalledCap[n,g,i]) > 3 else 0))
                        writer.writerow(my_string)
            f.close()

        if 'results_output_stor' in include_results:
            print("{hour}:{minute}:{second}: Writing storage investment decisions to results_output_stor.csv...".format(
                hour=datetime.now().strftime("%H"), minute=datetime.now().strftime("%M"), second=datetime.now().strftime("%S")))
            f = open(result_file_path + "/" + 'results_output_stor.csv', 'a', newline='')
            writer = csv.writer(f)
            if instance.CurrentPeriods[1]==1:
                writer.writerow(["Node","StorageType","Period","storPWInvCap_MW","storPWInstalledCap_MW","storENInvCap_MWh","storENInstalledCap_MWh","DiscountedInvestmentCostPWEN_EuroPerMWMWh","ExpectedAnnualDischargeVolume_GWh","ExpectedAnnualLossesChargeDischarge_GWh"])
            for (n,b) in instance.StoragesOfNode:
                for i in instance.CurrentPeriods:   
                    if i in updatePeriods or last_run == True:
                        writer.writerow([n,b,inv_per[int(i-1)],value(instance.storPWInvCap[n,b,i]),value(instance.storPWInstalledCap[n,b,i]),
                                            value(instance.storENInvCap[n,b,i]),value(instance.storENInstalledCap[n,b,i]),
                                            value(instance.discount_multiplier[i]*(instance.storPWInvCap[n,b,i]*instance.storPWInvCost[b,i] + instance.storENInvCap[n,b,i]*instance.storENInvCost[b,i])),
                                            value(sum(instance.branchProbab[w]*instance.seasScale[s]*instance.storDischarge[n,b,i,w,h]/1000 for (w,s,h) in instance.HoursAndSeasonOfBranch)),
                                            value(sum(instance.branchProbab[w]*instance.seasScale[s]*((1 - instance.storageDischargeEff[b])*instance.storDischarge[n,b,i,w,h] + (1 - instance.storageChargeEff[b])*instance.storCharge[n,b,i,w,h])/1000 for (w,s,h) in instance.HoursAndSeasonOfBranch))])
            f.close()

        if 'results_hydrogen_load_shed' in include_results:
            print("{hour}:{minute}:{second}: Writing hydrogen load shed results to 'results_hydrogen_load_shed'.csv...".format(
                hour=datetime.now().strftime("%H"), minute=datetime.now().strftime("%M"), second=datetime.now().strftime("%S")))
            f = open(result_file_path + "/" + 'results_hydrogen_load_shed.csv', 'a', newline='')
            writer = csv.writer(f)
            if instance.CurrentPeriods[1]==1:
                my_header = ["Node",'Period','Branch','Season','Hour','Demand met [tons]', 'Demand met SCALED [tons]', 'Demand shed [tons]', 'Demand shed SCALED [tons]']
                writer.writerow(my_header)
            for n in instance.Node:
                for i in instance.CurrentPeriods:   
                    if i in updatePeriods or last_run == True:
                        for (w,s,h) in instance.HoursAndSeasonOfBranch:
                                row = [n,inv_per[int(i)-1],w,s,h,
                                        value(instance.hydrogenDemandMet[n,i,w,h]),
                                        value(instance.seasScale[s] * instance.hydrogenDemandMet[n,i,w,h]),
                                        value(instance.hydrogenDemandShed[n,i,w,h]),
                                        value(instance.seasScale[s] * instance.hydrogenDemandShed[n,i,w,h])]
                                writer.writerow(row)
            f.close()

        if 'results_hydrogen_production_investments' in include_results:
            print("{hour}:{minute}:{second}: Writing hydrogen investment results to results_hydrogen_production_investments.csv...".format(
                hour=datetime.now().strftime("%H"), minute=datetime.now().strftime("%M"), second=datetime.now().strftime("%S")))
            f = open(result_file_path + "/" + 'results_hydrogen_production_investments.csv', 'a', newline='')
            writer = csv.writer(f)
            if instance.CurrentPeriods[1]==1:
                my_header = ["Node","Period","New electrolyzer capacity [MW]", "Total electrolyzer capacity [MW]", "New electrolyzer capacity [ton/h]", "Total electrolyzer capacity [ton/h]",
                            "Expected annual power usage [GWh]","Expected annual electrolyzer hydrogen production [ton]",
                            'Expected electrolyzer capacity factor']
                if REFORMER_HYDROGEN:
                    header.extend(['New Reformer capacity [ton/h]', 'Total Reformer capacity [ton/h]','Expected annual reformer hydrogen production [ton]'])
                writer.writerow(my_header)
            for n in instance.Node:
                for i in instance.CurrentPeriods:   
                    if i in updatePeriods or last_run == True:
                        if REFORMER_HYDROGEN:
                            ReformerCapBuilt = value(sum(instance.ReformerCapBuilt[n,p,i] for p in instance.ReformerPlants)/instance.hydrogenLHV_ton)
                            reformerCapTotal = value(sum(instance.ReformerTotalCap[n,p,i] for p in instance.ReformerPlants)/instance.hydrogenLHV_ton)
                            reformerExpectedProduction = value(sum(instance.seasScale[s] * instance.branchProbab[w] * instance.hydrogenProducedReformer_ton[n,p,i,w,h] for (w,s,h) in instance.HoursAndSeasonOfBranch for p in instance.ReformerPlants))
                        electrolyzerCapacity = value(instance.elyzerTotalCap[n,i] / instance.elyzerPowerConsumptionPerTon[i])
                        expectedElectrolyzerProduction = value(sum(instance.branchProbab[w] * instance.seasScale[s] * instance.hydrogenProducedElectro_ton[n,i,w,h] for (w,s,h) in instance.HoursAndSeasonOfBranch))
                        electrolyzerCapFactor = (expectedElectrolyzerProduction/(electrolyzerCapacity*8760) if electrolyzerCapacity > .001 else 0)
                        
                        row = [n,inv_per[int(i-1)],
                                        value(instance.elyzerCapBuilt[n,i]),
                                        value(instance.elyzerTotalCap[n,i]),
                                        value(instance.elyzerCapBuilt[n,i] / instance.elyzerPowerConsumptionPerTon[i]),
                                        electrolyzerCapacity,
                                        value(sum(instance.seasScale[s] * instance.branchProbab[w] * instance.powerForHydrogen[n,i,w,h] for (w,s,h) in instance.HoursAndSeasonOfBranch) / 1000),
                                        expectedElectrolyzerProduction,
                                        electrolyzerCapFactor]
                        if REFORMER_HYDROGEN:
                            row.extend([ReformerCapBuilt,
                                        reformerCapTotal,
                                        reformerExpectedProduction])
                        writer.writerow(row)
            f.close()

        if REFORMER_HYDROGEN and 'results_hydrogen_reformer_detailed_investments' in include_results:
            print("{hour}:{minute}:{second}: Writing detailed reformer investment results to results_hydrogen_reformer_detailed_investments.csv...".format(
                hour=datetime.now().strftime("%H"), minute=datetime.now().strftime("%M"), second=datetime.now().strftime("%S")))
            f = open(result_file_path + "/" + 'results_hydrogen_reformer_detailed_investments.csv', 'a', newline='')
            writer = csv.writer(f)
            if instance.CurrentPeriods[1]==1:
                my_header = ['Node','Reformer plant type','Period','New capacity [MW]','Total capacity [MW]','New capacity [ton/h]','Total capacity [ton/h]',
                                'Expected production [ton H2/year]', 'Expected capacity factor [%]', 'Expected emissions [tons CO2/year]', 'Expected electricity consumption [GWh]']
                writer.writerow(my_header)
            for n in instance.Node:
                for p in instance.ReformerPlants:
                    for i in instance.CurrentPeriods:   
                        if i in updatePeriods or last_run == True:
                            reformerCap = value(instance.ReformerTotalCap[n,p,i])
                            reformerProduction = value(sum(instance.branchProbab[w] * instance.seasScale[s] * instance.hydrogenProducedReformer_ton[n,p,i,w,h] for (w,s,h) in instance.HoursAndSeasonOfBranch))
                            capFactor = (reformerProduction / (8760*(reformerCap/value(instance.hydrogenLHV_ton))) if reformerCap > 1 else 0)
                            my_string = [n,p,inv_per[int(i)-1],
                                            value(instance.ReformerCapBuilt[n,p,i]),
                                            reformerCap,
                                            value(instance.ReformerCapBuilt[n,p,i]/instance.hydrogenLHV_ton),
                                            reformerCap/value(instance.hydrogenLHV_ton),
                                            reformerProduction,
                                            capFactor,
                                            reformerProduction * value(instance.ReformerEmissionFactor[p,i]),
                                            reformerProduction * value(instance.ReformerPlantElectricityUse[p,i]/1000)]
                            writer.writerow(my_string)
            f.close()

        if 'results_hydrogen_storage_investments' in include_results:
            print("{hour}:{minute}:{second}: Writing hydrogen storage investment results to results_hydrogen_storage_investments.csv...".format(
                hour=datetime.now().strftime("%H"), minute=datetime.now().strftime("%M"), second=datetime.now().strftime("%S")))
            f = open(result_file_path + "/" + 'results_hydrogen_storage_investments.csv', 'a', newline='')
            writer = csv.writer(f)
            my_header = ['Node','Period','New storage capacity [ton]','Total storage capacity [ton]', 'Discounted cost of new capacity [EUR]','Discounted total cost [EUR]']
            if instance.CurrentPeriods[1]==1:
                writer.writerow(my_header)
            for n in instance.Node:
                for i in instance.CurrentPeriods:   
                    if i in updatePeriods or last_run == True:
                        my_string = [n,inv_per[int(i)-1],
                                        value(instance.hydrogenStorageBuilt[n,i]),
                                        value(instance.hydrogenTotalStorage[n,i]),
                                        value(instance.hydrogenStorageBuilt[n,i] * instance.hydrogenStorageInvCost[i]),
                                        value(sum(instance.hydrogenStorageBuilt[n,j] * instance.hydrogenStorageInvCost[j] for j in instance.CurrentPeriods if j<=i))]

                        writer.writerow(my_string)
            f.close()

        if 'results_hydrogen_storage_operational' in include_results:
            print("{hour}:{minute}:{second}: Writing hydrogen storage operational results to results_hydrogen_storage_operational.csv...".format(
                hour=datetime.now().strftime("%H"), minute=datetime.now().strftime("%M"), second=datetime.now().strftime("%S")))
            f = open(result_file_path + "/" + 'results_hydrogen_storage_operational.csv', 'a', newline='')
            writer = csv.writer(f)
            my_header = ['Node','Period','Branch', 'Season',' Hour','Initial storage [ton]','Charge [ton]','Discharge [ton]','Final stored [ton]'] + [str(i) for i in range(1, value(len(instance.BranchPath))+1)]
            if instance.CurrentPeriods[1]==1:
                writer.writerow(my_header)
            for n in instance.Node:
                for i in instance.CurrentPeriods:   
                    if i in updatePeriods or last_run == True:
                        for (w,s,h) in instance.HoursAndSeasonOfBranch:
                                my_string= [n,inv_per[i-1], w, s, h]
                                if h in instance.FirstHoursOfRegSeason:
                                    my_string.extend([value(instance.hydrogenStorageOperational[n,i,w,h]+instance.hydrogenDischargeStorage[n,i,w,h]-instance.hydrogenChargeStorage[n,i,w,h])])
                                else:
                                    my_string.extend([value(instance.hydrogenStorageOperational[n,i,w,h-1])])
                                my_string.extend([value(instance.hydrogenChargeStorage[n,i,w,h]),
                                                    value(instance.hydrogenDischargeStorage[n,i,w,h]),
                                                    value(instance.hydrogenStorageOperational[n,i,w,h])])
                                for path in instance.BranchPath:
                                        if w in path:
                                            my_string.extend(['X'])
                                        else:
                                            my_string.extend([''])
                                writer.writerow(my_string)
            f.close()

        if 'results_hydrogen_production' in include_results:
            print("{hour}:{minute}:{second}: Writing hydrogen production results to results_hydrogen_production.csv...".format(
                hour=datetime.now().strftime("%H"), minute=datetime.now().strftime("%M"), second=datetime.now().strftime("%S")))
            f = open(result_file_path + '/' + 'results_hydrogen_production.csv', 'a', newline='')
            writer = csv.writer(f)
            if instance.CurrentPeriods[1]==1:
                my_header = ["Node", "Period", "Branch", "Season", "Hour", "Power for hydrogen [MWh]", "Electrolyzer production[ton]"]
                if REFORMER_HYDROGEN:
                    my_header.extend(['Reformer production [ton]','Total CO2 emission [ton CO2/ton H2]'])
                my_header.extend([str(i) for i in range(1, value(len(instance.BranchPath))+1)])
                writer.writerow(my_header)
            for n in instance.Node:
                for i in instance.CurrentPeriods:   
                    if i in updatePeriods or last_run == True:
                        for (w,s,h) in instance.HoursAndSeasonOfBranch:
                                my_string = [n, inv_per[int(i-1)], w, s, h,
                                            value(instance.powerForHydrogen[n,i,w,h]),
                                            value(instance.hydrogenProducedElectro_ton[n,i,w,h])]
                                # power_emissions_kg_per_MWh = calculatePowerEmissionIntensity(n,h,i,w)
                                if REFORMER_HYDROGEN:
                                    blue_h2_production_ton = value(sum(instance.hydrogenProducedReformer_ton[n,p,i,w,h] for p in instance.ReformerPlants))
                                    blue_h2_direct_emissions_ton = value(sum(instance.ReformerEmissionFactor[p,i] * instance.hydrogenProducedReformer_ton[n,p,i,w,h] for p in instance.ReformerPlants))
                                    # blue_h2_emissions_from_power_ton = power_emissions_kg_per_MWh * value(sum(instance.ReformerPlantElectricityUse[p,i] * instance.hydrogenProducedReformer_ton[n,p,i,w,h] for p in instance.ReformerPlants))
                                    blue_h2_emissions_from_power_ton = 0 # Emissions from use of power is 0, because marginal emissions from a sector (here, there power sector) which is already capped on emissions is 0.

                                    my_string.extend([blue_h2_production_ton])

                                    GREEN_HYDROGEN_production_ton = value(instance.hydrogenProducedElectro_ton[n,i,w,h])
                                    # GREEN_HYDROGEN_emissions_ton = power_emissions_ton_per_MWh * value(sum(instance.powerForHydrogen[n,j,i,w,h] for j in instance.CurrentPeriods if j<=i))
                                    GREEN_HYDROGEN_emissions_ton = 0 # Emissions from green H2 is 0, because marginal emissions from a sector (here, there power sector) which is already capped on emissions is 0.
                                    total_h2_production = blue_h2_production_ton + GREEN_HYDROGEN_production_ton
                                    if total_h2_production < .5:
                                        total_h2_emissions = 0
                                    else:
                                        total_h2_emissions = blue_h2_direct_emissions_ton + blue_h2_emissions_from_power_ton + GREEN_HYDROGEN_emissions_ton
                                    my_string.extend([total_h2_emissions])
                                for path in instance.BranchPath:
                                    if w in path:
                                        my_string.extend(['X'])
                                    else:
                                        my_string.extend([''])
                                writer.writerow(my_string)

        if 'results_hydrogen_use' in include_results:
            print("{hour}:{minute}:{second}: Writing hydrogen sales results to results_hydrogen_use.csv...".format(
                hour=datetime.now().strftime("%H"), minute=datetime.now().strftime("%M"), second=datetime.now().strftime("%S")))
            f = open(result_file_path + '/' + 'results_hydrogen_use.csv', 'a', newline='')
            if solver == 'Xpress':
                fError = open(result_file_path + "/" + "errorLog.log",'a')
            writer = csv.writer(f)
            if instance.CurrentPeriods[1]==1:
                my_header = ["Node", "Period", "Branch", "Season", "Hour", "Hydrogen produced [ton]" ,"Hydrogen stored [ton]", "Hydrogen withdrawn from storage [ton]", 'Initial storage level [ton]', 'Final storage level [ton]',"Hydrogen burned for power [ton]",
                            'Hydrogen exported [ton]', 'Hydrogen imported [ton]', 'Hydrogen demand met [ton]', 'Hydrogen demand shed [shed]','Hydrogen price [EUR/kg]'] + [str(i) for i in range(1, value(len(instance.BranchPath))+1)]
                writer.writerow(my_header)
            for n in instance.Node:
                for i in instance.CurrentPeriods:   
                    if i in updatePeriods or last_run == True:
                        for (w,s,h) in instance.HoursAndSeasonOfBranch:
                                my_string = [n, inv_per[int(i - 1)], w, s, h]
                                if REFORMER_HYDROGEN:
                                    my_string.extend([value(sum(instance.hydrogenProducedReformer_ton[n,p,i,w,h] for p in instance.ReformerPlants) + instance.hydrogenProducedElectro_ton[n,i,w,h])])
                                else:
                                    my_string.extend([value(instance.hydrogenProducedElectro_ton[n,i,w,h])])
                                my_string.extend([value(instance.hydrogenChargeStorage[n,i,w,h]),
                                                value(instance.hydrogenDischargeStorage[n,i,w,h]),
                                                value(instance.hydrogenStorageOperational[n,i,w,h] - instance.hydrogenChargeStorage[n,i,w,h] + instance.hydrogenDischargeStorage[n,i,w,h]),
                                                value(instance.hydrogenStorageOperational[n,i,w,h]),
                                                value(sum(instance.hydrogenForPower[g,n,i,w,h] for g in instance.HydrogenGenerators)),
                                                value(sum(instance.hydrogenSentPipeline[n,n2,i,w,h] for n2 in instance.HydrogenLinks[n])),
                                                value(sum(instance.hydrogenSentPipeline[n2,n,i,w,h] for n2 in instance.HydrogenLinks[n])),
                                                value(instance.hydrogenDemandMet[n,i,w,h]),
                                                value(instance.hydrogenDemandShed[n,i,w,h]),
                                                value(instance.dual[instance.hydrogen_flow_balance[n,i,w,h]]/(instance.discount_multiplier[i] * instance.operationalDiscountrate * instance.seasScale[s] * instance.branchProbab[w]) / 1000)])
                                for path in instance.BranchPath:
                                    if w in path:
                                        my_string.extend(['X'])
                                    else:
                                        my_string.extend([''])
                                        
                                writer.writerow(my_string)
            if solver == 'Xpress':
                fError.write('\n')
                fError.close()
        if 'results_hydrogen_pipeline_investments' in include_results:
            print("{hour}:{minute}:{second}: Writing hydrogen pipeline investment results to results_hydrogen_pipeline_investments.csv...".format(
                hour=datetime.now().strftime("%H"), minute=datetime.now().strftime("%M"), second=datetime.now().strftime("%S")))
            f = open(result_file_path + '/' + 'results_hydrogen_pipeline_investments.csv', 'a', newline='')
            writer = csv.writer(f)
            if instance.CurrentPeriods[1]==1:
                my_header = ["Between node", "And node", "Period", "Pipeline capacity built [ton/hr]", "Pipeline total capacity [ton/hr]",
                            "Discounted cost of (newly) built pipeline [EUR]", "Expected hydrogen transmission [tons]"]
                writer.writerow(my_header)
            for (n1,n2) in instance.HydrogenBidirectionPipelines:
                for i in instance.CurrentPeriods:   
                    if i in updatePeriods or last_run == True:
                        my_string = [n1, n2, inv_per[int(i-1)],
                                    value(instance.hydrogenPipelineBuilt[n1,n2,i]),
                                    value(instance.totalHydrogenPipelineCapacity[n1,n2,i]),
                                    value(instance.discount_multiplier[i] * (instance.hydrogenPipelineBuilt[n1,n2,i] * instance.hydrogenPipelineInvCost[n1,n2,i])),
                                    value(sum(instance.branchProbab[w]*instance.seasScale[s]*(instance.hydrogenSentPipeline[n1,n2,i,w,h] + instance.hydrogenSentPipeline[n2,n1,i,w,h])/1000 for (w,s,h) in instance.HoursAndSeasonOfBranch))]
                        writer.writerow(my_string)


        if 'results_hydrogen_pipeline_operational' in include_results:
            print("{hour}:{minute}:{second}: Writing hydrogen pipeline operational results to results_hydrogen_pipeline_operational.csv...".format(
                    hour=datetime.now().strftime("%H"), minute=datetime.now().strftime("%M"), second=datetime.now().strftime("%S")))
            f = open(result_file_path + '/' + 'results_hydrogen_pipeline_operational.csv', 'a', newline='')
            writer = csv.writer(f)
            if instance.CurrentPeriods[1]==1:
                my_header = ["From node", "To node", "Period", "Season", "Branch", "Hour", "Hydrogen sent [ton]", "Power consumed in each node for transmission (MWh)"] + [str(i) for i in range(1, value(len(instance.BranchPath))+1)]
                writer.writerow(my_header)
            for (n1,n2) in instance.AllowedHydrogenLinks:
                if (n1,n2) in instance.HydrogenBidirectionPipelines:
                    for i in instance.CurrentPeriods:   
                        if i in updatePeriods or last_run == True:
                            for (w,s,h) in instance.HoursAndSeasonOfBranch:
                                    my_string = [n1,n2,inv_per[int(i-1)],s,w,h,
                                                value(instance.hydrogenSentPipeline[n1,n2,i,w,h]),
                                                value(0.5*(instance.hydrogenSentPipeline[n1,n2,i,w,h] * instance.hydrogenPipelinePowerDemandPerTon[n1,n2]))]
                                    for path in instance.BranchPath:
                                        if w in path:
                                            my_string.extend(['X'])
                                        else:
                                            my_string.extend([''])
                                    writer.writerow(my_string)
                else:
                    for i in instance.CurrentPeriods:   
                        if i in updatePeriods or last_run == True:
                            for (w,s,h) in instance.HoursAndSeasonOfBranch:
                                    my_string = [n1,n2,inv_per[int(i-1)],s,w,h,
                                                value(instance.hydrogenSentPipeline[n1,n2,i,w,h]),
                                                value(0.5*(instance.hydrogenSentPipeline[n1,n2,i,w,h] * instance.hydrogenPipelinePowerDemandPerTon[n2,n1]))]
                                    for path in instance.BranchPath:
                                        if w in path:
                                            my_string.extend(['X'])
                                        else:
                                            my_string.extend([''])
                                    writer.writerow(my_string)

        if 'results_output_EuropeSummary' in include_results:
            print("{hour}:{minute}:{second}: Writing summary file to results_output_EuropeSummary.csv...".format(
                hour=datetime.now().strftime("%H"), minute=datetime.now().strftime("%M"), second=datetime.now().strftime("%S")))
            f = open(result_file_path + "/" + 'results_output_EuropeSummary.csv', 'a', newline='')
            writer = csv.writer(f)
            if instance.CurrentPeriods[1]==1:
                header = ["Period", "BranchPath", 'EmissionsFromPower_Ton', "CO2Price_EuroPerTon",
                      "CO2Cap_Ton"]
                writer.writerow(header)
            path=0
            for (w1,w2,w3,w4,w5,w6,w7,w8,w9,w10,w11,w12) in instance.BranchPath:
                path+=1
                for i in instance.CurrentPeriods:   
                    if i in updatePeriods or last_run == True:
                        branches = [w1,w2,w3,w4,w5,w6,w7,w8,w9,w10,w11,w12]
                        my_string=[inv_per[int(i-1)],path,
                                   value(instance.generatorEmissions[i,w1,w2,w3,w4,w5,w6,w7,w8,w9,w10,w11,w12]),
                                   -value(instance.dual[instance.emission_cap[i,w1,w2,w3,w4,w5,w6,w7,w8,w9,w10,w11,w12]]/(instance.discount_multiplier[i]*instance.operationalDiscountrate*co2_scale_factor)),
                                   value(instance.CO2cap[i]*1e6),
                        ]
                        writer.writerow(my_string)
            f.close()

        if 'results_output_Operational' in include_results:
            print("{hour}:{minute}:{second}: Writing operational results to results_output_Operational.csv...".format(
                hour=datetime.now().strftime("%H"), minute=datetime.now().strftime("%M"), second=datetime.now().strftime("%S")))
            f = open(result_file_path + "/" + 'results_output_Operational.csv', 'a', newline='')
            writer = csv.writer(f)
            if instance.CurrentPeriods[1]==1:
                
                my_header = ["Node","Period","Branch","Season","Hour","AllGen_MW","Load_MW","Net_load_MW"]
                for g in instance.Generator:
                    my_string = str(g)+"_MW"
                    my_header.append(my_string)
                if instance.CurrentPeriods[1]==1:
                    my_header.extend(["storCharge_MW","storDischarge_MW","storEnergyLevel_MWh","LossesChargeDischargeBleed_MW","FlowOut_MW","FlowIn_MW","LossesFlowIn_MW","LoadShed_MW","Price_EURperMWh","AvgCO2_kgCO2perMWh"])
                    my_header.extend([str(i) for i in range(1, value(len(instance.BranchPath))+1)])
                    writer.writerow(my_header)
            for n in instance.Node:
                for i in instance.CurrentPeriods:   
                    if i in updatePeriods or last_run == True:
                        for (w,s,h) in instance.HoursAndSeasonOfBranch:
                                my_string=[n,inv_per[int(i-1)],w,s,h,
                                value(sum(instance.genOperational[n,g,i,w,h] for g in instance.Generator if (n,g) in instance.GeneratorsOfNode)),
                                value(-instance.sload[n,i,w,h]),
                                value(-(instance.sload[n,i,w,h] - instance.loadShed[n,i,w,h] + sum(instance.storCharge[n,b,i,w,h] - instance.storageDischargeEff[b]*instance.storDischarge[n,b,i,w,h] for b in instance.Storage if (n,b) in instance.StoragesOfNode) +
                                sum(instance.transmissionOperational[n,link,i,w,h] - instance.lineEfficiency[link,n]*instance.transmissionOperational[link,n,i,w,h] for link in instance.NodesLinked[n])))]
                                for g in instance.Generator:
                                    if (n,g) in instance.GeneratorsOfNode:
                                        my_string.append(value(instance.genOperational[n,g,i,w,h]))
                                    else:
                                        my_string.append(0)
                                my_string.extend([value(sum(-instance.storCharge[n,b,i,w,h] for b in instance.Storage if (n,b) in instance.StoragesOfNode)),
                                                    value(sum(instance.storDischarge[n,b,i,w,h] for b in instance.Storage if (n,b) in instance.StoragesOfNode)),
                                                    value(sum(instance.storOperational[n,b,i,w,h] for b in instance.Storage if (n,b) in instance.StoragesOfNode)),
                                                    value(sum(-(1 - instance.storageDischargeEff[b])*instance.storDischarge[n,b,i,w,h] - (1 - instance.storageChargeEff[b])*instance.storCharge[n,b,i,w,h] - (1 - instance.storageBleedEff[b])*instance.storOperational[n,b,i,w,h] for b in instance.Storage if (n,b) in instance.StoragesOfNode)),
                                                    value(sum(-instance.transmissionOperational[n,link,i,w,h] for link in instance.NodesLinked[n])),
                                                    value(sum(instance.transmissionOperational[link,n,i,w,h] for link in instance.NodesLinked[n])),
                                                    value(sum(-(1 - instance.lineEfficiency[link,n])*instance.transmissionOperational[link,n,i,w,h] for link in instance.NodesLinked[n])),
                                                    value(instance.loadShed[n,i,w,h]),
                                                    value(instance.dual[instance.FlowBalance[n,i,w,h]]/(instance.discount_multiplier[i] * instance.operationalDiscountrate * instance.seasScale[s] * instance.branchProbab[w]))])
                                # if value(sum(instance.genOperational[n,g,i,w,h] for g in instance.Generator if (n,g) in instance.GeneratorsOfNode)) > 0:
                                #     my_string.extend([value(1000*sum(instance.genOperational[n,g,i,w,h]*instance.genCO2TypeFactor[g]*(3.6/instance.genEfficiency[g,i]) for g in instance.Generator if (n,g) in instance.GeneratorsOfNode)/sum(instance.genOperational[n,g,i,w,h] for g in instance.Generator if (n,g) in instance.GeneratorsOfNode))])
                                # else:
                                #     my_string.extend([0])
                                my_string.extend([calculatePowerEmissionIntensity(n,h,i,w)])
                                for path in instance.BranchPath:
                                    if w in path:
                                        my_string.extend(['X'])
                                    else:
                                        my_string.extend([''])
                                writer.writerow(my_string)
            f.close()
        if 'results_power_storage_operational' in include_results:
            print("{hour}:{minute}:{second}: Writing operational results to results_power_storage_operational.csv...".format(
                hour=datetime.now().strftime("%H"), minute=datetime.now().strftime("%M"), second=datetime.now().strftime("%S")))
            f = open(result_file_path + "/" + 'results_power_storage_operational.csv', 'a', newline='')
            writer = csv.writer(f)
            if instance.CurrentPeriods[1]==1:
                my_header = ["Node","Period","Branch","Season","Hour","All_InitialStorage", "All_FinalStorage"]
                for b in instance.Storage:
                    my_string = [str(b)+"_InitialStorage", str(b)+"_FinalStorage"]
                    my_header.extend(my_string)
                my_header.extend([str(i) for i in range(1, value(len(instance.BranchPath))+1)])
                writer.writerow(my_header)
            for n in instance.Node:
                for i in instance.CurrentPeriods:
                    if i in updatePeriods or last_run == True:
                        for (w,s,h) in instance.HoursAndSeasonOfBranch:
                                my_string=[n,inv_per[int(i-1)],w,s,h]

                                my_string.extend([value(sum(instance.storOperational[n,b,i,w,h]-instance.storageChargeEff[b]*instance.storCharge[n,b,i,w,h]+instance.storDischarge[n,b,i,w,h] for b in instance.Storage if (n,b) in instance.StoragesOfNode)),
                                                  value(sum(instance.storageBleedEff[b]*instance.storOperational[n,b,i,w,h] for b in instance.Storage if (n,b) in instance.StoragesOfNode))])
                                
                                for b in instance.Storage:
                                    if (n,b) in instance.StoragesOfNode:
                                        my_string.extend([value(instance.storOperational[n,b,i,w,h]-instance.storageChargeEff[b]*instance.storCharge[n,b,i,w,h]+instance.storDischarge[n,b,i,w,h]),
                                                  value(instance.storageBleedEff[b]*instance.storOperational[n,b,i,w,h])])
                                    else:
                                        my_string.extend([0,0])
                                for path in instance.BranchPath:
                                    if w in path:
                                        my_string.extend(['X'])
                                    else:
                                        my_string.extend([''])
                                writer.writerow(my_string)
            f.close()
        if 'results_output_transmission_operational' in include_results:
            print("{hour}:{minute}:{second}: Writing transmission operational decisions to results_output_transmission_operational.csv...".format(
                hour=datetime.now().strftime("%H"), minute=datetime.now().strftime("%M"), second=datetime.now().strftime("%S")))
            f = open(result_file_path + "/" + 'results_output_transmission_operational.csv', 'a', newline='')
            writer = csv.writer(f)
            if instance.CurrentPeriods[1]==1:
               writer.writerow(["FromNode","ToNode","Period","Season","Branch","Hour","TransmissionReceived_MW","Losses_MW"] + [str(i) for i in range(1, value(len(instance.BranchPath))+1)])
            for (n1,n2) in instance.DirectionalLink:
                for i in instance.CurrentPeriods:   
                    if i in updatePeriods or last_run == True:
                        for (w,s,h) in instance.HoursAndSeasonOfBranch:
                                transmissionSent = value(instance.transmissionOperational[n1,n2,i,w,h])
                                my_string = [n1,n2,inv_per[int(i-1)],s,w,h,
                                                    value(instance.lineEfficiency[n1,n2])*transmissionSent,
                                                    value((1 - instance.lineEfficiency[n1,n2]))*transmissionSent]
                                for path in instance.BranchPath:
                                    if w in path:
                                        my_string.extend(['X'])
                                    else:
                                        my_string.extend([''])
                                writer.writerow(my_string)
            f.close()

        if 'results_power_balance' in include_results:
            print(
                "{hour}:{minute}:{second}: Writing power balances to results_power_balance.csv...".format(
                    hour=datetime.now().strftime("%H"), minute=datetime.now().strftime("%M"),
                    second=datetime.now().strftime("%S")))
            f = open(result_file_path + "/" + 'results_power_balance.csv', 'a', newline='')
            writer = csv.writer(f)
            if instance.CurrentPeriods[1]==1:
                header = ["Node", "Period", "Season", "Hour", "Branch", "Available power [MWh]", "Power generation [MWh]", "Power curtailed [MWh]", "Power transmission in [MWh]","Power storage discharge [MWh]", "Power transmission out [MWh]", "Power storage charge [MWh]", "Power load [MWh]", "Power shed [MWh]"]
                header.append("Power for hydrogen [MWh]")
                header.extend([str(i) for i in range(1, value(len(instance.BranchPath))+1)])
                writer.writerow(header)
            for n in instance.Node:
                for i in instance.CurrentPeriods:   
                    if i in updatePeriods or last_run == True:
                        for (w,s,h) in instance.HoursAndSeasonOfBranch:
                                row = [n,inv_per[int(i-1)],s,h,w]
                                row.append(value(sum(instance.genCapAvail[n,g,w,h,i]*instance.genInstalledCap[n,g,i] for g in instance.Generator if (n,g) in instance.GeneratorsOfNode)))
                                row.append(value(sum(instance.genOperational[n,g,i,w,h] for g in instance.Generator if (n,g) in instance.GeneratorsOfNode)))
                                row.append(value(sum((instance.genCapAvail[n,g,w,h,i]*instance.genInstalledCap[n,g,i] - instance.genOperational[n,g,i,w,h]) for g in instance.Generator if (n,g) in instance.GeneratorsOfNode)))
                                row.append(value(sum(instance.lineEfficiency[link,n]*instance.transmissionOperational[link,n,i,w,h] for link in instance.NodesLinked[n])))
                                row.append(value(sum(instance.storageDischargeEff[b] * instance.storDischarge[n, b, i, w, h] for b in instance.Storage if (n, b) in instance.StoragesOfNode)))
                                row.append(value(sum(instance.transmissionOperational[n,link,i,w,h] for link in instance.NodesLinked[n])))
                                row.append(value(sum(instance.storCharge[n,b,i,w,h] for b in instance.Storage if (n,b) in instance.StoragesOfNode)))
                                row.append(value(instance.sload[n,i,w,h]))
                                row.append(value(instance.loadShed[n,i,w,h]))
                                row.append(value(instance.powerForHydrogen[n,i,w,h]))
                                for path in instance.BranchPath:
                                    if w in path:
                                        row.extend(['X'])
                                    else:
                                        row.extend([''])
                                writer.writerow(row)
            f.close()

        if 'results_output_curtailed_prod' in include_results:
            print("{hour}:{minute}:{second}: Writing curtailed power to results_output_curtailed_prod.csv...".format(
                hour=datetime.now().strftime("%H"), minute=datetime.now().strftime("%M"), second=datetime.now().strftime("%S")))
            f = open(result_file_path + "/" + 'results_output_curtailed_prod.csv', 'a', newline='')
            writer = csv.writer(f)
            if instance.CurrentPeriods[1]==1:
                writer.writerow(["Node","RESGeneratorType","Period","ExpectedAnnualCurtailment_GWh", "Expected total available power_GWh", "Expected annual curtailment ratio of total capacity_%"])
            for t in instance.Technology:
                if t == 'Hydro_ror' or t == 'Wind_onshr' or t == 'Wind_offshr_grounded' or t == 'Wind_offshr_floating' or t == 'Solar':
                    for (n,g) in instance.GeneratorsOfNode:
                        if (t,g) in instance.GeneratorsOfTechnology:
                            for i in instance.CurrentPeriods:
                                if i in updatePeriods or last_run == True:
                                    curtailedPower = value(sum(instance.branchProbab[w]*instance.seasScale[s]*(instance.genCapAvail[n,g,w,h,i]*instance.genInstalledCap[n,g,i] - instance.genOperational[n,g,i,w,h])/1000 for (w,s,h) in instance.HoursAndSeasonOfBranch))
                                    totalPowerProduction = value(sum(instance.branchProbab[w]*instance.seasScale[s]*(instance.genCapAvail[n,g,w,h,i]*instance.genInstalledCap[n,g,i])/1000 for (w,s,h) in instance.HoursAndSeasonOfBranch))
                                    row = [n,g,inv_per[int(i-1)], curtailedPower, totalPowerProduction]
                                    if totalPowerProduction > 0:
                                        row.append(curtailedPower/totalPowerProduction*100)
                                    else:
                                        row.append(0)
                                    writer.writerow(row)
            f.close()
    finally:
        endReporting = timeEnd = datetime.now()

    if PICKLE_INSTANCE:
        print(("{hour}:{minute}:{second}: Saving instance").format(
        hour=datetime.now().strftime("%H"), minute = datetime.now().strftime("%M"), second=datetime.now().strftime("%S")))
        pickle_start = datetime.now()
        start = time.time()
        picklestring = 'instance' + name + '.pkl'
        if USE_TEMP_DIR:
            picklestring = './Finished instances/' + name + '.pkl'
        with open(picklestring, mode='wb') as file:
            cloudpickle.dump(instance, file)
        end = time.time()
        pickle_end = datetime.now()
        print("Pickling instance took [sec]:")
        print(str(end - start))

    if 'time_usage' in include_results:
        print("{hour}:{minute}:{second}: Writing time usage to time_usage.csv...".format(
            hour=datetime.now().strftime("%H"), minute=datetime.now().strftime("%M"), second=datetime.now().strftime("%S")))
    
        f = open(result_file_path + "/" + 'time_usage.csv', 'w', newline='')
        timeFrmt = "%H:%M:%S"
        dateFrmt = "%d.%m.%Y"
        timeDeltaFrmt = "{H}:{M}:{S}"
        writer = csv.writer(f)
        if (timeEnd - timeStart).days > 0:
            writer.writerow(["Process",
                             "Time started [HH:MM:SS]",
                             "Time ended [HH:MM:SS]",
                             "Time spent [HH:MM:SS]",
                             "Date started [DD.MM.YYYY]",
                             "Date finished [DD.MM.YYYY]"])
            writer.writerow(["Overall",
                             timeStart.strftime(timeFrmt),
                             timeEnd.strftime(timeFrmt),
                             strfdelta(timeEnd-timeStart,timeDeltaFrmt),
                             timeStart.strftime(dateFrmt),
                             timeEnd.strftime(dateFrmt)])
            writer.writerow(["Declaring and reading sets & parameters",
                             timeStart.strftime(timeFrmt),
                             stopReading.strftime(timeFrmt),
                             strfdelta(stopReading-timeStart,timeDeltaFrmt),
                             timeStart.strftime(dateFrmt),
                             stopReading.strftime(dateFrmt)])
            writer.writerow(["Declaring variables & constraints",
                             startConstraints.strftime(timeFrmt),
                             stopConstraints.strftime(timeFrmt),
                             strfdelta(stopConstraints-startConstraints,timeDeltaFrmt),
                             startConstraints.strftime(dateFrmt),
                             stopConstraints.strftime(dateFrmt)])
            writer.writerow(["Building model",
                             startBuild.strftime(timeFrmt),
                             endBuild.strftime(timeFrmt),
                             strfdelta(endBuild-startBuild,timeDeltaFrmt),
                             startBuild.strftime(dateFrmt),
                             endBuild.strftime(dateFrmt)])
            writer.writerow(["Optimizing model",
                             startOptimization.strftime(timeFrmt),
                             endOptimization.strftime(timeFrmt),
                             strfdelta(endOptimization-startOptimization,timeDeltaFrmt),
                             startOptimization.strftime(dateFrmt),
                             endOptimization.strftime(dateFrmt)])
            if PICKLE_INSTANCE:
                writer.writerow(["Saving instance",
                             pickle_start.strftime(timeFrmt),
                             pickle_end.strftime(timeFrmt),
                             strfdelta(pickle_end-pickle_start,timeDeltaFrmt),
                             pickle_start.strftime(dateFrmt),
                             pickle_end.strftime(dateFrmt)])
            writer.writerow(["Reporting results",
                             StartReporting.strftime(timeFrmt),
                             endReporting.strftime(timeFrmt),
                             strfdelta(endReporting-StartReporting,timeDeltaFrmt),
                             StartReporting.strftime(dateFrmt),
                             endReporting.strftime(dateFrmt)])
        else:
            writer.writerow(["Process",
                             "Time started [HH:MM:SS]",
                             "Time ended [HH:MM:SS]",
                             "Time spent [HH:MM:SS]"])
            writer.writerow(["Overall",
                             timeStart.strftime(timeFrmt),
                             timeEnd.strftime(timeFrmt),
                             strfdelta(timeEnd - timeStart, timeDeltaFrmt)])
            writer.writerow(["Declaring and reading sets & parameters",
                             timeStart.strftime(timeFrmt),
                             stopReading.strftime(timeFrmt),
                             strfdelta(stopReading - timeStart, timeDeltaFrmt)])
            writer.writerow(["Declaring variables & constraints",
                             startConstraints.strftime(timeFrmt),
                             stopConstraints.strftime(timeFrmt),
                             strfdelta(stopConstraints - startConstraints, timeDeltaFrmt)])
            writer.writerow(["Building model",
                             startBuild.strftime(timeFrmt),
                             endBuild.strftime(timeFrmt),
                             strfdelta(endBuild - startBuild, timeDeltaFrmt)])
            writer.writerow(["Optimizing model",
                             startOptimization.strftime(timeFrmt),
                             endOptimization.strftime(timeFrmt),
                             strfdelta(endOptimization - startOptimization,
                                       timeDeltaFrmt)])
            if PICKLE_INSTANCE:
                writer.writerow(["Saving instance",
                             pickle_start.strftime(timeFrmt),
                             pickle_end.strftime(timeFrmt),
                             strfdelta(pickle_end-pickle_start,timeDeltaFrmt),
                             pickle_start.strftime(dateFrmt),
                             pickle_end.strftime(dateFrmt)])
            writer.writerow(["Reporting results",
                             StartReporting.strftime(timeFrmt),
                             endReporting.strftime(timeFrmt),
                             strfdelta(endReporting - StartReporting, timeDeltaFrmt)])
        f.close()


    if 'numerics_info' in include_results:
        print("{hour}:{minute}:{second}: Writing numerical information to numerics_info.csv...".format(
            hour=datetime.now().strftime("%H"), minute=datetime.now().strftime("%M"), second=datetime.now().strftime("%S")))
        f = open(result_file_path + "/" + 'numerics_info.csv', 'w', newline='')
        writer = csv.writer(f)
        header = ["Number of variables", "Number of constraints"]# "Maximum constraint matrix coefficient", "Minimum constraint matrix coefficient", "Maximum RHX", "Minimum RHS"]
        writer.writerow(header)
        my_str = [instance.nvariables(), instance.nconstraints()]
        writer.writerow(my_str)
        f.close()



    print("{hour}:{minute}:{second} Finished writing results to files.".format(
        hour=datetime.now().strftime("%H"), minute=datetime.now().strftime("%M"), second=datetime.now().strftime("%S")))

    del results, instance, model