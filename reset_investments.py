import pandas as pd
import numpy as np
import os
import openpyxl
from openpyxl import load_workbook

def reset_investment_files(path):
    sheets = [
        ('Generator.xlsx','GeneratorInv','A,B,C'),
        ('Transmission.xlsx','TransmissionInv','A,B,C'),
        ('Storage.xlsx','StoragePWInv','A,B,C'),
        ('Storage.xlsx','StorageENInv','A,B,C'),
        ('Hydrogen.xlsx','ElyzerInv','A,B'),
        ('Hydrogen.xlsx','PipelineInv','A,B,C'),
        ('Hydrogen.xlsx','StorageInv','A,B'),
        ('Hydrogen.xlsx','ElyzerInv','A,B'),
        ('Generator.xlsx','RESGeneratorInv','A,B,C'),
        #('Industry.xlsx','SteelInv','A,B,C'),
        #('Industry.xlsx','CementInv','A,B,C'),
        #('Industry.xlsx','AmmoniaInv','A,B,C'),
        ('Hydrogen.xlsx','ReformerInv','A,B,C')
    ]
    
    for (file, sheet_name, cols) in sheets:
        file_path = os.path.join(path, file)
        df = pd.read_excel(file_path, sheet_name=sheet_name, usecols=cols, skiprows=2)
        df['Capacity'] = 0

        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
            if sheet_name in workbook.sheetnames:
                # Get the existing sheet object
                ws = workbook[sheet_name]
                # Remove the existing sheet
                workbook.remove(ws)
                # Create a new sheet with the same name
                sheet = workbook.create_sheet(sheet_name)
            else:
                # If the sheet does not exist, create a new one
                sheet = workbook.create_sheet(sheet_name)
        else:
            # If the file does not exist, create a new workbook and sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = sheet_name

        # Set headers for rows 1 and 2
        sheet['A1'] = 'Source: EMPIRE'
        sheet['A2'] = 'Description: Investment decisions for each period'

        # Write DataFrame headers to row 3
        for col, header in enumerate(df.columns, start=1):
            sheet.cell(row=3, column=col, value=header)

        # Write DataFrame data starting from row 4
        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(df, index=False, header=False)
        for r_idx, row in enumerate(rows, start=4):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

        # Save the workbook
        workbook.save(file_path)

